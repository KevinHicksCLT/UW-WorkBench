#!/usr/bin/env python
"""Update a row in the Transformation Bridge Defect Tracker.

Sets Agent Status (and optionally Review Status / Pending Review Comments) for a
defect ID, AND applies the matching state colour as an explicit cell fill so the
colour holds even on rows outside the conditional-formatting range (C5:C55).
Conditional formatting and dropdown validations are preserved by openpyxl on save.

Usage:
  python update_defect.py --id VS-01 --status Complete
  python update_defect.py --id T-05 --status Complete --comment "AC verified: ..."
  python update_defect.py --id E-02 --status "In Progress"
  python update_defect.py --id VS-01 --review "Needs Changes" --comment "..."
  python update_defect.py --list            # show all IDs + current status

Run from repo root (or pass --file). Close the xlsx in Excel first or save fails.
"""
import argparse, sys, os
import openpyxl
from openpyxl.styles import PatternFill

DEFAULT_FILE = "documents/Transformation-Bridge-Defect-Tracker.xlsx"
SHEET = "Defect Tracker"
HEADER_ROW = 4
COL = {"id": 1, "agent_status": 3, "review_status": 4, "comments": 5}  # A,C,D,E

AGENT_STATES = {
    "Not Started": "FCE4D6",
    "In Progress": "FFEB9C",
    "Blocked":     "FFC7CE",
    "Complete":    "C6EFCE",
}
REVIEW_STATES = {
    "Pending Review": "FFEB9C",
    "Approved":       "C6EFCE",
    "Rejected":       "FFC7CE",
    "Needs Changes":  "FCE4D6",
}


def fill(hex6):
    return PatternFill(patternType="solid", fgColor="FF" + hex6)


def find_row(ws, did):
    for r in range(HEADER_ROW + 1, ws.max_row + 1):
        if str(ws.cell(r, COL["id"]).value).strip() == did:
            return r
    return None


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--id")
    p.add_argument("--status", choices=list(AGENT_STATES))
    p.add_argument("--review", choices=list(REVIEW_STATES))
    p.add_argument("--comment")
    p.add_argument("--file", default=DEFAULT_FILE)
    p.add_argument("--list", action="store_true")
    a = p.parse_args()

    if not os.path.exists(a.file):
        sys.exit("File not found: %s (run from repo root or pass --file)" % a.file)
    wb = openpyxl.load_workbook(a.file)
    ws = wb[SHEET]

    if a.list:
        for r in range(HEADER_ROW + 1, ws.max_row + 1):
            i = ws.cell(r, COL["id"]).value
            if i:
                print("%-8s %-12s | %s" % (i, ws.cell(r, COL["agent_status"]).value,
                                           ws.cell(r, COL["review_status"]).value))
        return

    if not a.id:
        sys.exit("--id required (or use --list)")
    r = find_row(ws, a.id)
    if not r:
        sys.exit("ID not found: %s" % a.id)

    if a.status:
        c = ws.cell(r, COL["agent_status"]); c.value = a.status; c.fill = fill(AGENT_STATES[a.status])
        print("Agent Status -> %s" % a.status)
    if a.review:
        c = ws.cell(r, COL["review_status"]); c.value = a.review; c.fill = fill(REVIEW_STATES[a.review])
        print("Review Status -> %s" % a.review)
    if a.comment:
        ws.cell(r, COL["comments"]).value = a.comment
        print("Comment updated")

    wb.save(a.file)
    print("Saved row %d (%s)" % (r, a.id))


if __name__ == "__main__":
    main()
