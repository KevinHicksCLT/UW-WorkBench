# -*- coding: utf-8 -*-
import json, re, sys
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from collections import Counter, defaultdict
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
O = "documents/value-streams/_master_build"
def nk(s): return re.sub(r"[^a-z0-9]+", " ", (s or "").lower()).strip()
F = r"C:\Users\xando\AppData\Local\Temp\claude\C--Users-xando-Code-transform-platform\e7779fde-4039-436b-94a9-de90aeff60c5\tasks\wagtt3bgd.output"
res = json.load(open(F, encoding="utf-8"))["result"]
groups = res["groups"]
l4_by_l3 = json.load(open(f"{O}/l4_by_l3.json", encoding="utf-8"))
pools = json.load(open(f"{O}/pools_by_division.json", encoding="utf-8"))
key_l1 = {k: g["l1"] for k, g in l4_by_l3.items()}
expected_l4 = {(g["l2"], g["l3"], l4) for g in l4_by_l3.values() for l4 in g["l4s"]}

GENERIC = {"initiate & capture","analyze & assess","execute & process","validate & control","finalize & handoff","define requirements","develop approach"}
rows = []; covered_l4 = set(); flagged = 0; echo = 0; genname = 0
for g in groups:
    k = g["l3key"]; l1 = key_l1.get(k, "")
    parts = [x.strip() for x in k.split("||")]; l2 = parts[0]; l3 = parts[1] if len(parts) > 1 else ""
    pool = pools.get(l2, {})
    for l4g in g.get("l4groups", []):
        l4 = l4g["l4"]; covered_l4.add((l2, l3, l4))
        for t in l4g.get("tasks", []):
            name = t.get("name", ""); deliv = t.get("deliverable", "")
            if nk(deliv).endswith("results"): echo += 1
            if nk(name) in GENERIC: genname += 1
            okv = t.get("ok", True)
            if not okv: flagged += 1
            rows.append({
                "l1": l1, "l2": l2, "l3": l3, "l4": l4, "name": name,
                "owner": t.get("owner", ""), "participants": "; ".join(t.get("participants", [])),
                "app": t.get("application", ""), "deliv": deliv,
                "checklist": " | ".join(t.get("checklist", [])), "auto": t.get("automatability", ""),
                "standards": pool.get("standards", ""), "regs": pool.get("regulations", ""),
                "metrics": "; ".join(pool.get("metrics", [])[:12]),
                "ok": okv, "issues": t.get("issues", ""),
            })
rows.sort(key=lambda r: (r["l1"], r["l2"], r["l3"], r["l4"], r["name"]))

# QA scorecard (six questions)
dist = Counter()
for r in rows:
    q1 = bool(r["owner"]); q2 = bool(r["deliv"]); q3 = bool(r["name"])
    q4 = bool(r["app"]); q5 = bool(r["checklist"] or r["standards"] or r["regs"]); q6 = bool(r["metrics"])
    r["comp"] = sum([q1, q2, q3, q4, q5, q6]); dist[r["comp"]] += 1

# workbook
wb = openpyxl.Workbook()
YEL = PatternFill("solid", fgColor="FFF3B0"); HDR = PatternFill("solid", fgColor="1F3864"); HF = Font(color="FFFFFF", bold=True)
FLG = PatternFill("solid", fgColor="F8D7DA")
COLS = ["L1 Segment","L2 Division","L3 Process","L4 Sub-Process","L5 Task","isTask","Owner / Lead Role",
        "Roles (Participants)","Applications / SoR","Deliverable","Checklist","Automatability",
        "Standards","Regulations","Metrics","Complete /6","Flag","Issues"]
ws = wb.active; ws.title = "MASTER - L5 (rebuilt)"
ws.append(COLS)
for j in range(1, len(COLS) + 1):
    c = ws.cell(1, j); c.fill = HDR; c.font = HF; c.alignment = Alignment(vertical="center", wrap_text=True)
for r in rows:
    rec = [r["l1"], r["l2"], r["l3"], r["l4"], r["name"], "Y", r["owner"], r["participants"], r["app"],
           r["deliv"], r["checklist"], r["auto"], r["standards"], r["regs"], r["metrics"], f"{r['comp']}/6",
           "" if r["ok"] else "FLAG", r["issues"]]
    ws.append(rec); rr = ws.max_row
    if not r["app"]: ws.cell(rr, 9).fill = YEL
    if not r["ok"]: ws.cell(rr, 17).fill = FLG
ws.freeze_panes = "E2"
for i, w in enumerate([13,22,26,28,40,5,22,30,22,34,50,12,20,24,30,8,6,30], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{len(rows)+1}"

sc = wb.create_sheet("Scorecard")
sc.append(["FINAL REBUILD SCORECARD"]); sc.cell(1,1).font = Font(bold=True, size=13, color="1F3864")
sc.append([]); sc.append(["L5 tasks (rebuilt)", len(rows)])
sc.append(["Verified clean", res["verifiedOk"]]); sc.append(["Flagged for review", flagged])
sc.append(["Template/echo defects remaining", echo + genname, "(was 402 in Bridge)"])
sc.append([]); sc.append(["Structure", ""])
sc.append(["L1 Segment", len({r['l1'] for r in rows})]); sc.append(["L2 Division", len({(r['l1'],r['l2']) for r in rows})])
sc.append(["L3 Process", len({(r['l1'],r['l2'],r['l3']) for r in rows})])
sc.append(["L4 Sub-Process (covered)", len(covered_l4), f"of {len(expected_l4)} expected"])
sc.append(["L5 Task", len(rows)])
sc.append([]); sc.append(["Completeness /6", "tasks"])
for k in range(6, -1, -1): sc.append([f"{k}/6", dist.get(k, 0)])
sc.append([]); sc.append(["Note", "5/6 are tasks with no system-of-record (manual) - app left blank/yellow by design"])
for i, w in enumerate([34, 12, 26], 1): sc.column_dimensions[get_column_letter(i)].width = w

# missing L4 coverage
miss = sorted(expected_l4 - covered_l4)
mc = wb.create_sheet("Coverage gaps")
mc.append(["L4 sub-processes with no rebuilt tasks", len(miss)]); mc.cell(1,1).font = Font(bold=True)
mc.append(["L2 Division","L3 Process","L4 Sub-Process"])
for l2, l3, l4 in miss: mc.append([l2, l3, l4])
for i, w in enumerate([22, 28, 34], 1): mc.column_dimensions[get_column_letter(i)].width = w

rd = wb.create_sheet("README")
for ln in [
 ["ABC Insurance - MASTER (L5 rebuilt, grounded)"],
 [""],
 ["APPROACH", "Kept the L1-L4 skeleton you approved (3 / 17 / 135 / 867). Rebuilt L5 as logical atomic tasks (~4-5 per L4) grounded in each value stream's real doer-roles and systems of record; adversarially verified."],
 ["WHAT IS GROUNDED", "Owner/Participants = real doer roles for that value stream (no executives). Application = real system of record. Standards/Regulations/Metrics = real, from production, at value-stream grain."],
 ["WHAT IS DERIVED", "L5 task names, deliverables, and checklists are generated to be specific and atomic, then verified for logic. Tasks marked FLAG (red) failed verification - review them."],
 ["YELLOW", "Application blank = manual task with no system of record (legitimate); review if a system applies."],
 ["NOT TOUCHED", "Database is unchanged. This sheet is the staging/QA artifact for sign-off before any load."],
] : rd.append(ln)
rd.column_dimensions["A"].width = 22; rd.column_dimensions["B"].width = 120
rd.cell(1,1).font = Font(bold=True, size=14, color="1F3864")

out = "documents/value-streams/ABC-Insurance-Operating-Model-MASTER-v7-REBUILT.xlsx"
wb.save(out)
print("SAVED", out)
print("L5 tasks", len(rows), "| flagged", flagged, "| template/echo remaining", echo + genname)
print("L4 covered", len(covered_l4), "of", len(expected_l4), "| missing", len(miss))
print("completeness", dict(sorted(dist.items())))
