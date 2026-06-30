# -*- coding: utf-8 -*-
# Accountability-aware owner matching: pick owner at the RIGHT level for the task's altitude
# (execute -> IC/specialist; manage/plan/govern -> manager/head; approve -> senior), responsibility-matched.
import json, re, sys
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from collections import Counter
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
O = "documents/value-streams/_master_build"
def Lj(n): return json.load(open(f"{O}/{n}", encoding="utf-8"))
def toks(s):
    STOP = set("the a an and or of to for in on with by from into against per as at is are be vs via using use".split())
    return {w for w in re.findall(r"[a-z0-9]+", (s or "").lower()) if w not in STOP and len(w) > 2}

tba = Lj("tasks_by_area.json"); rref = Lj("role_ref_by_division.json")
apool = Lj("app_pool_by_division.json"); l4_by_l3 = Lj("l4_by_l3.json")
key_l1 = {k: g["l1"] for k, g in l4_by_l3.items()}
LLM = {}
try:
    res = json.load(open(r"C:\Users\xando\AppData\Local\Temp\claude\C--Users-xando-Code-transform-platform\e7779fde-4039-436b-94a9-de90aeff60c5\tasks\wf1jc1kvj.output", encoding="utf-8"))["result"]
    for g in res["groups"]: LLM[g["l3key"]] = {(l4g["l4"]): l4g["tasks"] for l4g in g.get("l4groups", [])}
except Exception as e:
    print("LLM apps overlay skipped:", e)

def rank(role, level):
    n = role.lower(); lv = (level or "").lower()
    if any(w in n for w in ["chief ", "head of", " officer", "general counsel", "president", "appointed actuary"]) or lv == "executive": return 3
    if "director" in n or lv == "leadership": return 3
    if "manager" in n or n.endswith(" lead") or " lead " in n or "principal" in n or lv == "manager": return 2
    if any(w in n for w in ["senior", "lead "]) : return 1
    # specialists/ICs
    return 0

GOVERN = re.compile(r"\bbudget|operating plan|business plan|\bplan\b|planning|resource|staffing|capacity plan|headcount|governance|oversight|strateg|roadmap|prioriti|allocat|operating model|charter|set (the )?standard|define (the )?(approach|framework|policy)|manage the|coordinate the program|portfolio管理|workforce|target operating")
APPROVE = re.compile(r"approve|sign.?off|authoriz|certif|attest|final (decision|approval)|go.?no.?go|board (approval|review)|adjudicat(e|ion) authority")
def altitude(name):
    n = name.lower()
    if APPROVE.search(n): return "approve", 3
    if GOVERN.search(n): return "govern", 2
    return "execute", 0

def pick(name, deliverable, l4, roles):
    alt, target = altitude(name)
    ttok = toks(name + " " + deliverable + " " + l4)
    scored = []
    for r in roles:
        rk = rank(r["role"], r.get("level"))
        ov = len(ttok & toks(r["role"] + " " + " ".join(r.get("does", [])) + " " + (r.get("family") or "")))
        d = abs(rk - target)
        fit = 5 if d == 0 else (2 if d == 1 else -5)
        # never put an executive on pure execution
        if alt == "execute" and rk >= 3: fit -= 6
        # execution should not be owned by a manager when an IC fits
        score = ov * 2 + fit
        scored.append((score, rk, ov, r["role"]))
    if not scored: return "", []
    scored.sort(key=lambda x: (-x[0], -x[2]))
    owner = scored[0][3]; owner_rk = scored[0][1]
    # participants: doers (rank <= owner, prefer one level below) with real overlap, excluding owner
    docand = [s for s in scored if s[3] != owner]
    # prefer lower-rank executors with overlap
    docand.sort(key=lambda x: (-(x[2]), x[1]))
    participants = []
    for s in docand:
        if s[3] not in participants: participants.append(s[3])
        if len(participants) >= 2: break
    return owner, participants

# apps (reuse rules from v8) constrained to division pool
APP_RULES = [
 (r"go.?no.?go|launch|release|deploy|go.?live|readiness|sprint|backlog|user story|cut.?over|pipeline build", ["Jira", "Jira Align / Rally", "Azure DevOps"]),
 (r"change ticket|incident|service request|change advisory|cab\b|problem record", ["ServiceNow ITSM"]),
 (r"ledger|journal|financial close|reconcil|general ledger|accru|trial balance", ["SAP S/4HANA (GL)", "BlackLine (Close)"]),
 (r"budget|forecast|fp&a|variance|planning cycle|operating plan", ["Oracle EPM", "OneStream (FP&A)", "Anaplan (Planning)"]),
 (r"invoice|billing|premium bill|collection|receivable|payment|dunning|cash appl", ["Guidewire BillingCenter", "Payment Gateway"]),
 (r"\bclaim|fnol|adjud|settle|salvage|adjust|loss notice|coverage determination", ["Guidewire ClaimCenter", "OpenText Document Mgmt"]),
 (r"fraud|siu\b", ["FRISS Fraud Detection", "Guidewire ClaimCenter"]),
 (r"underwrit|submission|quote|bind|appetite|clearance|risk assessment|pric(e|ing) indication", ["Guidewire PolicyCenter", "Earnix Rating & Pricing", "Broker & Agent Portal"]),
 (r"policy admin|endorse|renew|issuance|servicing|in.?force", ["Guidewire PolicyCenter", "Policy Administration Platform"]),
 (r"reserv|experience study|mortality|morbidity|lapse|loss cost|loss development|capital model|solvency|actuar|assumption", ["SAS Actuarial Platform", "FIS Prophet", "Milliman Arius (Reserving)"]),
 (r"cat |catastrophe|exposure model|pml", ["Moody's RMS (Cat Model)", "Catastrophe Data Provider"]),
 (r"reinsur|cede|treaty|retrocess|bordereau|recover", ["Sapiens ReinsurancePro", "Reinsurer Exchange"]),
 (r"\bdata\b|extract|dataset|etl|ingest|dashboard|report build|data quality|master data", ["Snowflake Data Cloud", "Databricks Lakehouse", "Power BI"]),
 (r"\bmodel|machine learning|\bml\b|train|inference|drift|mlops|feature", ["Azure ML / MLflow", "Databricks Lakehouse"]),
 (r"investment|portfolio|trade|asset alloc|aladdin|securities", ["BlackRock Aladdin", "Bloomberg Terminal"]),
 (r"cyber|vulnerab|threat|siem|patch|malware|soc\b", ["Splunk SIEM", "CrowdStrike Falcon", "Qualys / Tenable (Vuln)"]),
 (r"identity|access review|provision|\biam\b|privileged|certif|entitlement", ["SailPoint IGA", "Okta / IAM Platform", "CyberArk (PAM)"]),
 (r"recruit|hire|onboard|talent acquis|requisition|candidate", ["Greenhouse (Recruiting)", "Workday HCM"]),
 (r"payroll|compensation|benefit|employee|learning|training|performance review", ["Workday HCM", "ADP Payroll", "Cornerstone LMS"]),
 (r"contract|legal|litigat|counsel|agreement|nda|dispute", ["iManage (Legal Docs)", "Icertis CLM (Contracts)", "DocuSign"]),
 (r"compliance|regulat|filing|exam|\baudit|control|grc|\baml\b|sanction|conduct", ["Archer GRC", "AuditBoard", "NICE Actimize (AML)", "Regulatory Filing Portal (SERFF)"]),
 (r"campaign|marketing|brand|content|\blead\b", ["Marketo (Marketing)", "Salesforce FSC"]),
 (r"producer|agent|broker|distribut|appointment|licens", ["Duck Creek Distribution", "NIPR (Producer Licensing)", "Broker & Agent Portal", "Salesforce FSC"]),
 (r"customer|complaint|contact center|call\b|service level|csat", ["Salesforce FSC", "Genesys Cloud (Contact Center)", "Customer Self-Service Portal"]),
 (r"vendor|procure|third.?party|sourcing", ["Coupa / SAP Ariba (Procurement)", "Archer GRC"]),
]
DOC_RE = re.compile(r"document|draft|report|memo|publish|procedure|minutes|present|communicat|sign.?off|approve|recommend|charter|\bplan\b|budget")
def assign_apps(name, l4, pool_names):
    text = (name + " " + l4).lower(); picks = []
    for rx, apps in APP_RULES:
        if re.search(rx, text):
            for a in apps:
                if a in pool_names and a not in picks: picks.append(a)
    if DOC_RE.search(text):
        for a in ("Microsoft 365 / SharePoint", "Confluence"):
            if a in pool_names and a not in picks: picks.append(a)
    if not picks:
        sors = [a for a in pool_names if a not in ("Microsoft 365 / SharePoint", "Microsoft Excel", "Microsoft Teams / Email")]
        picks = sors[:1] or ["Microsoft 365 / SharePoint"]
    return picks[0], picks[1:3]

COLS = ["L1 Segment","L2 Division","L3 Process","L4 Sub-Process","L5 Task","isTask","Owner / Lead Role",
        "Roles (Participants)","Primary App / SoR","Supporting Apps","Deliverable","Checklist","Automatability","Altitude"]
rows = []; altc = Counter()
for k, area in tba.items():
    parts = [x.strip() for x in k.split("||")]; l2 = parts[0]; l3 = parts[1] if len(parts) > 1 else ""
    l1 = key_l1.get(k, ""); roles = rref.get(l2, []); pool_names = [a["name"] for a in apool.get(l2, [])]
    llm_apps = LLM.get(k, {})
    for g in area.get("l4groups", []):
        l4 = g["l4"]
        for t in g.get("tasks", []):
            name = t["name"]; deliv = t.get("deliverable", "")
            owner, pcs = pick(name, deliv, l4, roles)
            alt, _ = altitude(name); altc[alt] += 1
            # apps: prefer LLM-verified apps for this task if present
            la = next((x for x in llm_apps.get(l4, []) if x.get("name") == name), None)
            if la and la.get("primary_app"):
                primary = la["primary_app"]; supp = la.get("supporting_apps", [])[:2]
            else:
                primary, supp = assign_apps(name, l4, pool_names)
            rows.append([l1, l2, l3, l4, name, "Y", owner, "; ".join(pcs), primary, "; ".join(supp),
                         deliv, " | ".join(t.get("checklist", [])), t.get("automatability", ""), alt])
rows.sort(key=lambda r: (r[0], r[1], r[2], r[3], r[4]))

wb = openpyxl.Workbook(); ws = wb.active; ws.title = "MASTER - L5 (fixed)"
HDR = PatternFill("solid", fgColor="1F3864"); HF = Font(color="FFFFFF", bold=True); YEL = PatternFill("solid", fgColor="FFF3B0")
ws.append(COLS)
for j in range(1, len(COLS) + 1):
    c = ws.cell(1, j); c.fill = HDR; c.font = HF; c.alignment = Alignment(vertical="center", wrap_text=True)
for r in rows:
    ws.append(r); rr = ws.max_row
    if not r[6]: ws.cell(rr, 7).fill = YEL
    if not r[8]: ws.cell(rr, 9).fill = YEL
ws.freeze_panes = "E2"
for i, w in enumerate([13,22,26,28,40,5,24,30,22,28,34,50,12,9], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{len(rows)+1}"
out = "documents/value-streams/ABC-Insurance-Operating-Model-MASTER-v9-FIXED.xlsx"
wb.save(out)
print("SAVED", out, "| tasks", len(rows), "| altitude", dict(altc))
# spot checks
def show(sub):
    for r in rows:
        if sub.lower() in r[4].lower():
            print(f"  [{r[1]}] {r[4]}\n    owner: {r[6]} | participants: {r[7]}"); return
show("department operating budget"); show("Extract in-force and claims data"); show("go/no-go review and record launch")
