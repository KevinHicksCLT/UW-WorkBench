# -*- coding: utf-8 -*-
# Bridge-sourced MASTER: tasks + associations pulled from AI Transformation Bridge Input.
# Value streams kept (all L2 Divisions). Where Bridge has no data -> blank + yellow.
import json, re, sys
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
O = "documents/value-streams/_master_build"
def L(n): return json.load(open(f"{O}/{n}", encoding="utf-8"))
def nk(s): return re.sub(r"[^a-z0-9]+", " ", (s or "").lower()).strip()
def cap(s, n=600): s = s or ""; return s if len(s) <= n else s[:n] + " ..."

v007 = L("wb_v007.json"); bl4 = L("wb_bridge_l4.json"); bl5 = L("wb_bridge_l5.json")
stds = L("wb_standards.json"); bext = L("wb_ext.json"); bmet = L("wb_metrics.json")
db_app = L("db_app.json"); db_role = L("db_role.json"); appcat = L("wb_appcat.json")

# L2 Division -> production/Bridge Value Stream crosswalk (Bridge L2VS names == these)
XWALK = {
 "actuarial": ["Actuarial Pricing, Reserving & Capital Modeling"],
 "business operations": ["Policy Administration & Servicing", "Service Operations, Incident & Production Support"],
 "call center": ["Customer Service, Complaints & Experience"],
 "claims": ["Claims Intake-to-Settlement", "Claims Recoveries & Subrogation", "Life & Annuity Claims"],
 "corporate functions operations": ["Enterprise Strategy & Portfolio Management", "Third-Party & Vendor Management"],
 "cybersecurity iam": ["Cybersecurity, Identity & Resilience"],
 "data ai": ["Data, Analytics & AI Management", "MLOps / ML Lifecycle Management", "AIOps & Intelligent Operations"],
 "finance investments": ["Finance, Treasury & Capital Management", "Investment & Asset Management", "FinOps / Cloud Financial Management", "Billing, Collections & Receivables"],
 "human resources talent": ["Talent & Workforce Management"],
 "legal corporate governance": ["Legal, Governance & Privacy Management"],
 "product delivery": ["Product & Proposition Management", "Technology Delivery & Change"],
 "program management office": ["Change Management & Adoption"],
 "reinsurance": ["Reinsurance & Retrocession Management"],
 "risk compliance audit": ["Risk, Compliance & Regulatory Management", "Audit & Assurance"],
 "sales distribution marketing": ["Distribution & Channel Management", "Marketing, Growth & Customer Insights"],
 "technology engineering": ["Technology Strategy, Architecture & Delivery"],
 "underwriting": ["Submission-to-Bind / Underwriting", "Delegated Authority Management"],
}
STD_X = {"actuarial":"actuarial","claims":"claims","underwriting":"underwriting","technology engineering":"engineering",
 "data ai":"data","finance investments":"finance","human resources talent":"hr","risk compliance audit":"compliance",
 "legal corporate governance":"legal","cybersecurity iam":"cyber","product delivery":"pmo","program management office":"pmo",
 "business operations":"operations","call center":"operations","corporate functions operations":"operations"}

div_display = {}; div_l1 = {}
for r in v007:
    d = nk(r.get("L2 Division"))
    if d and d not in div_display:
        div_display[d] = r.get("L2 Division"); div_l1[d] = r.get("L1 Segment")
rev = {}
for dnk, vss in XWALK.items():
    for vs in vss: rev.setdefault(nk(vs), dnk)
std_map = {nk(r.get("Department / Team")): {"count": r.get("Standards Count"), "owner": r.get("Primary Owner")} for r in stds if r.get("Department / Team")}
bl4i = {(nk(r.get("L3 Process Area")), nk(r.get("L4 Sub-Process"))): r for r in bl4}
met_by = {}
for r in bmet:
    met_by.setdefault(nk(r.get("L2 Value Stream")), []).append(r.get("Metric Name"))
def domain_seg(dom):
    d = nk(dom)
    if "technology" in d or "data" in d: return "IT"
    if "core insurance" in d or "operations customer" in d or "product growth" in d or d == "operations control": return "Core Business"
    return "Corporate Functions"
def ext_match(l2, l3):
    hits = []
    for r in bext:
        rv = nk(r.get("Related Value Stream")) + " " + nk(r.get("Division / Function"))
        if (nk(l2) and nk(l2) in rv) or (nk(l3) and nk(l3) in rv):
            hits.append(f"{r.get('External Party Type')}/{r.get('External Role')}")
    return "; ".join(dict.fromkeys(hits))[:250]

COLS = ["L1 Segment","L2 Division","L3 Process","L4 Sub-Process","L5 Step / Task","Step Description","isTask",
        "Owner / Lead Role","Roles (Participants)","Applications / SoR","Deliverables","Standards",
        "Regulations","Metrics","External Interactions","Checklist","Automatability",
        "Bridge Step ID","Match Level","Sources","Assoc Status"]
YELLOW = list(range(8, 18))  # Owner..Automatability
master = []; covered = set()
for t in bl5:
    vs = t.get("L2 Value Stream"); dnk = rev.get(nk(vs))
    if dnk: l1 = div_l1.get(dnk, domain_seg(t.get("Value Stream Domain"))); l2 = div_display.get(dnk, vs); covered.add(dnk)
    else: l1 = domain_seg(t.get("Value Stream Domain")); l2 = vs + "  [unmapped VS]"
    l3 = t.get("L3 Process Area"); l4 = t.get("L4 Sub-Process"); l5 = t.get("L5 Process Step Name")
    owner = t.get("Step Lead(s)")
    parts = t.get("Supporting Participants") or (bl4i.get((nk(l3), nk(l4))) or {}).get("Key Roles", "")
    deliv = t.get("Key Outputs")
    std = ""; kw = STD_X.get(dnk or "")
    if kw:
        for dk, dv in std_map.items():
            if kw in dk: std = f"{dv['count']} standards - {dv['owner']}"; break
    metrics = "; ".join([m for m in dict.fromkeys(met_by.get(nk(vs), [])) if m])
    ext = t.get("External Participants") or ext_match(l2, l3)
    src = ["Bridge"] + (["Std"] if std else [])
    rec = [l1, l2, l3, l4, l5, cap(t.get("Step Description"), 300), "Y",
           owner, cap(parts, 300), "", cap(deliv, 300), std, "", cap(metrics, 400), cap(ext, 250), "", "",
           t.get("L5 Step ID"), "Bridge-L5", "; ".join(src), ""]
    nfill = sum(1 for i in [7, 8, 10, 11, 13, 14] if rec[i])
    rec[20] = "Associated" if nfill >= 4 else ("Partial" if nfill >= 1 else "GAP")
    master.append(rec)
# keep all value streams: add placeholder rows for locked divisions Bridge did not cover
kept_gap = []
for dnk, disp in div_display.items():
    if dnk not in covered:
        rec = [div_l1.get(dnk), disp, "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "none", "GAP (value stream kept - no Bridge tasks)", "GAP"]
        master.append(rec); kept_gap.append(disp)
# sort
master.sort(key=lambda r: (str(r[0]), str(r[1]), str(r[2]), str(r[3]), str(r[4])))

# ---------------- workbook ----------------
wb = openpyxl.Workbook()
YEL = PatternFill("solid", fgColor="FFF3B0"); HDR = PatternFill("solid", fgColor="1F3864"); HF = Font(color="FFFFFF", bold=True, size=11)
GAPF = PatternFill("solid", fgColor="F8D7DA"); PARTF = PatternFill("solid", fgColor="FFF8E1"); OKF = PatternFill("solid", fgColor="E8F5E9")
ws = wb.active; ws.title = "MASTER - L5 (Bridge)"; ws.append(COLS)
for j in range(1, len(COLS) + 1):
    c = ws.cell(1, j); c.fill = HDR; c.font = HF; c.alignment = Alignment(vertical="center", wrap_text=True)
colfill = {c: 0 for c in COLS}
for rec in master:
    ws.append(rec); r = ws.max_row
    for j in YELLOW:
        if not ws.cell(r, j).value: ws.cell(r, j).fill = YEL
    ws.cell(r, 21).fill = {"GAP": GAPF, "Partial": PARTF, "Associated": OKF}[rec[20]]
    for i, c in enumerate(COLS):
        if rec[i]: colfill[c] += 1
ws.freeze_panes = "F2"
for i, w in enumerate([13,22,26,28,34,40,6,22,34,14,30,22,12,34,26,12,11,14,11,16,12], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{len(master)+1}"

# ---------------- Bridge Gap Analysis ----------------
ga = wb.create_sheet("Bridge Gap Analysis")
def H(s): ga.append([s]); ga.cell(ga.max_row, 1).font = Font(bold=True, color="1F3864", size=12)
bridge_tasks = len(bl5); mapped = sum(1 for t in bl5 if rev.get(nk(t.get("L2 Value Stream"))))
unmapped_vs = sorted({t.get("L2 Value Stream") for t in bl5 if not rev.get(nk(t.get("L2 Value Stream")))})
H("BRIDGE GAP ANALYSIS  (master vs AI Transformation Bridge Input)")
ga.append([])
H("1) Bridge inventory (the source of truth)")
ga.append(["Bridge L5 tasks", bridge_tasks]); ga.append(["Bridge L4 sub-processes", len(bl4)])
ga.append(["Bridge L2 value streams", len({nk(t.get('L2 Value Stream')) for t in bl5})])
ga.append(["Bridge domains", len({nk(t.get('Value Stream Domain')) for t in bl5})])
ga.append([])
H("2) Mapping Bridge -> kept value streams")
ga.append(["Bridge tasks mapped to a locked L2 Division", mapped])
ga.append(["Bridge tasks under UNMAPPED value streams", bridge_tasks - mapped])
ga.append(["Unmapped Bridge value streams (kept as-is):"])
for v in unmapped_vs: ga.append(["", v])
ga.append([])
H("3) Value-stream coverage (all 17 kept divisions)")
ga.append(["L2 Division", "Bridge tasks", "Status"])
from collections import Counter
tc = Counter(r[1] for r in master if r[18] == "Bridge-L5")
for dnk, disp in sorted(div_display.items(), key=lambda x: x[1]):
    n = tc.get(disp, 0); ga.append([disp, n, "covered" if n else "GAP - no Bridge tasks (kept, yellow)"])
ga.append([])
H("4) Association coverage in Bridge (what Bridge can / cannot supply)")
ga.append(["Column", "Filled", "of " + str(len(master)), "Note"])
notes = {"Owner / Lead Role":"Bridge Step Lead(s)","Roles (Participants)":"Bridge Supporting + L4 Key Roles (the doers, not execs)",
 "Applications / SoR":"Bridge App-Usage is placeholder (4 rows) -> BLANK/yellow","Deliverables":"Bridge Key Outputs",
 "Standards":"Bridge Standards Index by domain","Regulations":"Bridge has NO per-task regs -> BLANK/yellow",
 "Metrics":"Bridge Value Stream Metrics","External Interactions":"Bridge External Participants",
 "Checklist":"Bridge has NO checklists -> BLANK/yellow","Automatability":"not scored yet -> BLANK/yellow"}
for c in COLS[7:17]: ga.append([c, colfill[c], "", notes.get(c, "")])
ga.append([])
H("5) What the locked spine had that Bridge does NOT (the gaps)")
ga.append(["Locked spine L5 steps (synthetic)", len(v007)])
ga.append(["Bridge curated L5 tasks", bridge_tasks])
ga.append(["=> synthetic L5 steps with no Bridge backing", len(v007) - bridge_tasks])
ga.append(["Locked L2 Divisions with NO Bridge tasks", "; ".join(kept_gap) or "(none)"])
ga.append(["Columns Bridge cannot fill (all yellow)", "Applications/SoR, Checklist, Regulations, Automatability"])
for i, w in enumerate([46, 14, 14, 60], 1): ga.column_dimensions[get_column_letter(i)].width = w

# ---------------- README ----------------
rd = wb.create_sheet("README & Legend")
for ln in [
 ["ABC Insurance - MASTER (Bridge-sourced)"],
 [""],
 ["SOURCE OF TRUTH", "AI Transformation Bridge Input.xlsx - tasks and associations are pulled from Bridge for correctness."],
 ["GRAIN", f"One row per Bridge L5 task ({len(bl5)} tasks) + placeholder rows for value streams Bridge does not cover."],
 ["ROLES FIX", "Roles are now the people DOING the task (Bridge Step Lead + Supporting Participants / L4 Key Roles) - NOT value-stream execs."],
 ["VALUE STREAMS", "All value streams kept. Bridge tasks grouped under locked L2 Divisions via the L2<->VS crosswalk; uncovered divisions kept as yellow placeholders."],
 ["REMOVED", "Inputs and Outputs columns removed (per request)."],
 [""],
 ["COLOR KEY"],
 ["  Yellow cell", "Bridge has no data here -> BLANK and flagged (Applications, Checklist, Regulations, Automatability, plus any empty Bridge field)."],
 ["  Status Associated/Partial/GAP", "green / amber / red by # of Bridge associations present."],
 [""],
 ["NOT IN BRIDGE (always yellow)", "Per-task Applications/SoR, Checklists, Regulations, Automatability. These need a separate source or manual entry."],
 ["NEXT", "Review the Bridge Gap Analysis tab; confirm the L2<->VS crosswalk; decide how to source Applications / Checklists / Regulations."],
] : rd.append(ln)
rd.column_dimensions["A"].width = 34; rd.column_dimensions["B"].width = 115
rd.cell(1, 1).font = Font(bold=True, size=14, color="1F3864"); rd.cell(10, 1).fill = YEL

def cat(name, rows, cols, keys):
    s = wb.create_sheet(name); s.append(cols)
    for j in range(1, len(cols) + 1): s.cell(1, j).fill = HDR; s.cell(1, j).font = HF
    for r in rows: s.append([r.get(k, "") for k in keys])
    s.freeze_panes = "A2"
    for j, w in enumerate([34, 20, 18, 18, 22, 30], 1):
        if j <= len(cols): s.column_dimensions[get_column_letter(j)].width = w
    s.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"
cat("Catalog - Applications (Bridge)", appcat, ["App ID","Application / Tool","Vendor","Category","Primary Domain","SoR?"], ["App ID","Application / Tool","Vendor (example)","Category","Primary Domain","System-of-Record?"])
cat("Catalog - Standards (Bridge)", stds, ["Department / Team","Count","Charter","Owner"], ["Department / Team","Standards Count","Charter Included","Primary Owner"])
cat("Catalog - Roles (DB ref)", db_role, ["Role","Family","Level","Domain","Primary VS"], ["name","fam","lvl","dom","pvs"])

out = "documents/value-streams/ABC-Insurance-Operating-Model-MASTER-v4-Bridge.xlsx"
wb.save(out)
print("SAVED", out)
print("master rows", len(master), "(bridge tasks", bridge_tasks, "+ kept-gap divisions", len(kept_gap), ")")
print("mapped", mapped, "unmapped VS", len(unmapped_vs), unmapped_vs)
print("kept-gap divisions:", kept_gap)
print("col fill:", {c: colfill[c] for c in COLS[7:17]})
