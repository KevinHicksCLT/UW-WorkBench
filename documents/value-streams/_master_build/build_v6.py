# -*- coding: utf-8 -*-
import json, sys
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from collections import Counter, defaultdict
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
O = "documents/value-streams/_master_build"
F = r"C:\Users\xando\AppData\Local\Temp\claude\C--Users-xando-Code-transform-platform\2a55fb36-bab5-4963-a470-37a0a5b6a5d0\tasks\wy29knv6z.output"
res = json.load(open(F, encoding="utf-8"))["result"]
defects = res["defects"]
json.dump(defects, open(f"{O}/defects.json", "w", encoding="utf-8"), ensure_ascii=False)

bl5 = json.load(open(f"{O}/wb_bridge_l5.json", encoding="utf-8"))
GENERIC = {"initiate & capture","analyze & assess","execute & process","validate & control","finalize & handoff",
 "define requirements","develop approach","detect and log","execute and process","document findings","build and configure","review and approve"}
def nk(s): return (s or "").strip().lower()
scaffold = set()
for t in bl5:
    if nk(t.get("Key Outputs")).endswith("results") or nk(t.get("L5 Process Step Name")) in GENERIC:
        scaffold.add(t.get("L5 Step ID"))
defbyid = Counter(d["id"] for d in defects)

# open v5, enrich, save v6
SRC = "documents/value-streams/ABC-Insurance-Operating-Model-MASTER-v5-QA.xlsx"
wb = openpyxl.load_workbook(SRC)
ws = wb["MASTER + QA"]
hdr = [c.value for c in ws[1]]
idcol = hdr.index("Bridge Step ID") + 1
ncol = ws.max_column
ws.cell(1, ncol + 1, "Scaffold?"); ws.cell(1, ncol + 2, "Defects (n)")
HDR = PatternFill("solid", fgColor="1F3864"); HF = Font(color="FFFFFF", bold=True)
SCF = PatternFill("solid", fgColor="F8D7DA"); REALF = PatternFill("solid", fgColor="E8F5E9")
for j in (ncol + 1, ncol + 2): ws.cell(1, j).fill = HDR; ws.cell(1, j).font = HF
sc_n = 0
for r in range(2, ws.max_row + 1):
    tid = ws.cell(r, idcol).value
    isscf = tid in scaffold
    ws.cell(r, ncol + 1, "TEMPLATE" if isscf else "real")
    ws.cell(r, ncol + 1).fill = SCF if isscf else REALF
    n = defbyid.get(tid, 0); ws.cell(r, ncol + 2, n if n else "")
    if isscf: sc_n += 1
ws.column_dimensions[get_column_letter(ncol + 1)].width = 11
ws.column_dimensions[get_column_letter(ncol + 2)].width = 10

# QA Defects tab
qd = wb.create_sheet("QA Defects (verified)")
cols = ["Severity", "Value Stream", "Task ID", "Task Name", "Field", "Problem", "Suggestion"]
qd.append(cols)
for j in range(1, len(cols) + 1): qd.cell(1, j).fill = HDR; qd.cell(1, j).font = HF
sev_rank = {"high": 0, "medium": 1, "low": 2}
SEVF = {"high": PatternFill("solid", fgColor="F8D7DA"), "medium": PatternFill("solid", fgColor="FDE2C8"), "low": PatternFill("solid", fgColor="FFF8E1")}
for d in sorted(defects, key=lambda d: (sev_rank.get(d["severity"], 9), d["vs"], d["id"])):
    qd.append([d["severity"], d["vs"], d["id"], d.get("name", ""), d["field"], d["problem"], d["suggestion"]])
    qd.cell(qd.max_row, 1).fill = SEVF.get(d["severity"], PatternFill())
qd.freeze_panes = "A2"
for i, w in enumerate([9, 26, 16, 30, 12, 70, 60], 1): qd.column_dimensions[get_column_letter(i)].width = w
qd.auto_filter.ref = f"A1:G{len(defects)+1}"

# Defect Summary tab
ds = wb.create_sheet("Defect Summary")
def H(s): ds.append([s]); ds.cell(ds.max_row, 1).font = Font(bold=True, color="1F3864", size=12)
H("VERIFIED DATA-QUALITY AUDIT — SUMMARY"); ds.append([])
H("Scaffold vs real (Bridge 711 tasks)")
ds.append(["Template / scaffold tasks", sc_n, f"{round(100*sc_n/711)}%"])
ds.append(["Genuinely curated tasks", 711 - sc_n, f"{round(100*(711-sc_n)/711)}%"])
ds.append(["  - deliverable echoes 'X results'", sum(1 for t in bl5 if nk(t.get('Key Outputs')).endswith('results'))])
ds.append(["  - generic stage-template name", sum(1 for t in bl5 if nk(t.get('L5 Process Step Name')) in GENERIC)])
ds.append([])
H("Verified defects (adversarially confirmed)")
ds.append(["Total", res["count"]])
ds.append([]); ds.append(["By field", ""])
for k, v in sorted(res["byField"].items(), key=lambda x: -x[1]): ds.append([k, v])
ds.append([]); ds.append(["By severity", ""])
for k in ("high", "medium", "low"): ds.append([k, res["bySeverity"].get(k, 0)])
ds.append([]); ds.append(["By value stream (top 15)", ""])
vc = Counter(d["vs"] for d in defects)
for vs, n in vc.most_common(15): ds.append([vs, n])
for i, w in enumerate([40, 12, 8], 1): ds.column_dimensions[get_column_letter(i)].width = w

out = "documents/value-streams/ABC-Insurance-Operating-Model-MASTER-v6-AUDITED.xlsx"
wb.save(out)
print("SAVED", out)
print("scaffold/template tasks flagged:", sc_n, "of 711 | curated:", 711 - sc_n)
print("verified defects written:", len(defects), "| byField", res["byField"], "| bySeverity", res["bySeverity"])
print("top VS by defects:", vc.most_common(6))
