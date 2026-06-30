# -*- coding: utf-8 -*-
# Six-questions QA harness + coverage/reconciliation, on the Bridge-sourced master.
import json, re, sys
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from collections import Counter, defaultdict
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
O = "documents/value-streams/_master_build"
def L(n): return json.load(open(f"{O}/{n}", encoding="utf-8"))
def nk(s): return re.sub(r"[^a-z0-9]+", " ", (s or "").lower()).strip()
def cap(s, n=600): s = s or ""; return s if len(s) <= n else s[:n] + " ..."

v007 = L("wb_v007.json"); bl4 = L("wb_bridge_l4.json"); bl5 = L("wb_bridge_l5.json")
stds = L("wb_standards.json"); bext = L("wb_ext.json"); bmet = L("wb_metrics.json"); db_role = L("db_role.json")
VF = L("vs_fill.json"); vsapps = VF["vsapps"]; vsreg_n = VF["vsreg_n"]; vsreg_titles = VF["vsreg_titles"]

XWALK = {
 "actuarial": ["Actuarial Pricing, Reserving & Capital Modeling"],
 "business operations": ["Policy Administration & Servicing", "Service Operations, Incident & Production Support"],
 "call center": ["Customer Service, Complaints & Experience"],
 "claims": ["Claims Intake-to-Settlement", "Claims Recoveries & Subrogation", "Life & Annuity Claims"],
 "corporate functions operations": ["Enterprise Strategy & Portfolio Management", "Third-Party & Vendor Management"],
 "cybersecurity iam": ["Cybersecurity, Identity & Resilience"],
 "data ai": ["Data, Analytics & AI Management", "MLOps / ML Lifecycle Management", "AIOps & Intelligent Operations"],
 "finance investments": ["Finance, Treasury & Capital Management", "Investment & Asset Management", "FinOps / Cloud Financial Management", "Billing, Collections & Receivables"],
 "human resources talent": ["Talent & Workforce Management"],
 "legal corporate governance": ["Legal, Governance & Privacy Management"],
 "product delivery": ["Product & Proposition Management", "Technology Delivery & Change"],
 "program management office": ["Change Management & Adoption"],
 "reinsurance": ["Reinsurance & Retrocession Management"],
 "risk compliance audit": ["Risk, Compliance & Regulatory Management", "Audit & Assurance"],
 "sales distribution marketing": ["Distribution & Channel Management", "Marketing, Growth & Customer Insights"],
 "technology engineering": ["Technology Strategy, Architecture & Delivery"],
 "underwriting": ["Submission-to-Bind / Underwriting", "Delegated Authority Management"],
}
STD_X = {"actuarial":"actuarial","claims":"claims","underwriting":"underwriting","technology engineering":"engineering",
 "data ai":"data","finance investments":"finance","human resources talent":"hr","risk compliance audit":"compliance",
 "legal corporate governance":"legal","cybersecurity iam":"cyber","product delivery":"pmo","program management office":"pmo",
 "business operations":"operations","call center":"operations","corporate functions operations":"operations"}

div_display, div_l1 = {}, {}
for r in v007:
    d = nk(r.get("L2 Division"))
    if d and d not in div_display: div_display[d] = r.get("L2 Division"); div_l1[d] = r.get("L1 Segment")
rev = {}
for dnk, vss in XWALK.items():
    for vs in vss: rev.setdefault(nk(vs), dnk)
std_map = {nk(r.get("Department / Team")): {"count": r.get("Standards Count"), "owner": r.get("Primary Owner")} for r in stds if r.get("Department / Team")}
bl4i = {(nk(r.get("L3 Process Area")), nk(r.get("L4 Sub-Process"))): r for r in bl4}
met_by = defaultdict(list)
for r in bmet: met_by[nk(r.get("L2 Value Stream"))].append(r.get("Metric Name"))
# role -> level catalog (for exec detection)
rolelvl = {nk(r.get("name")): nk(r.get("lvl")) for r in db_role if r.get("name")}
EXEC = {"executive", "leadership"}
def role_tokens(*cells):
    out = []
    for c in cells:
        for p in re.split(r"[;,/]", c or ""):
            p = p.strip()
            if p: out.append(p)
    return out
def exec_hits(*cells):
    hits = []
    for r in role_tokens(*cells):
        lv = rolelvl.get(nk(r), "")
        if lv in EXEC or nk(r).startswith("chief ") or nk(r).startswith("head of") or "officer" in nk(r):
            hits.append(r)
    return list(dict.fromkeys(hits))
def domain_seg(dom):
    d = nk(dom)
    if "technology" in d or "data" in d: return "IT"
    if "core insurance" in d or "operations customer" in d or "product growth" in d or d == "operations control": return "Core Business"
    return "Corporate Functions"
def ext_match(l2, l3):
    hits = [f"{r.get('External Party Type')}/{r.get('External Role')}" for r in bext
            if (nk(l2) and nk(l2) in nk(r.get("Related Value Stream")) + " " + nk(r.get("Division / Function")))
            or (nk(l3) and nk(l3) in nk(r.get("Related Value Stream")) + " " + nk(r.get("Division / Function")))]
    return "; ".join(dict.fromkeys(hits))[:250]

# ---- build rows + QA
BASECOLS = ["L1 Segment","L2 Division","L3 Process","L4 Sub-Process","L5 Step / Task","Step Description","isTask",
            "Owner / Lead Role","Roles (Participants)","Applications / SoR","Deliverables","Standards",
            "Regulations","Metrics","External Interactions","Checklist","Automatability","Bridge Step ID"]
QACOLS = ["Q1 who+lead","Q2 deliverable","Q3 how","Q4 where(app)","Q5 controls","Q6 metrics","Complete /6","Exec flags (R1)","QA Status"]
COLS = BASECOLS + QACOLS
rows = []; audit_rows = []
dist = Counter(); r1count = 0; statc = Counter()
for t in bl5:
    vs = t.get("L2 Value Stream"); dnk = rev.get(nk(vs))
    if dnk: l1 = div_l1.get(dnk, domain_seg(t.get("Value Stream Domain"))); l2 = div_display.get(dnk, vs)
    else: l1 = domain_seg(t.get("Value Stream Domain")); l2 = vs + "  [unmapped VS]"
    l3 = t.get("L3 Process Area"); l4 = t.get("L4 Sub-Process"); l5 = t.get("L5 Process Step Name")
    owner = t.get("Step Lead(s)"); parts = t.get("Supporting Participants") or (bl4i.get((nk(l3), nk(l4))) or {}).get("Key Roles", "")
    deliv = t.get("Key Outputs"); desc = t.get("Step Description")
    std = ""; kw = STD_X.get(dnk or "")
    if kw:
        for dk, dv in std_map.items():
            if kw in dk: std = f"{dv['count']} standards - {dv['owner']}"; break
    metrics = "; ".join([m for m in dict.fromkeys(met_by.get(nk(vs), [])) if m])
    ext = t.get("External Participants") or ext_match(l2, l3)
    apps = "; ".join(vsapps.get(nk(vs), []))  # grounded VS-grain systems of record (DB)
    rn = vsreg_n.get(nk(vs), 0)
    regs = (f"{rn} requirements - {vsreg_titles.get(nk(vs), '')}".strip(" -") if rn > 0 else "")
    base = [l1, l2, l3, l4, l5, cap(desc, 300), "Y", owner, cap(parts, 300), cap(apps, 260), cap(deliv, 300), std, cap(regs, 300), cap(metrics, 400), cap(ext, 250), "", "", t.get("L5 Step ID")]
    q1 = bool(owner) and bool(parts); q2 = bool(deliv); q3 = bool(l5) and bool(desc)
    q4 = bool(apps); q5 = bool(std or regs)  # controls = standards or regulations (checklist still blank)
    q6 = bool(metrics)
    eh = exec_hits(owner, parts); r1 = bool(eh); r1count += 1 if r1 else 0
    comp = sum([q1, q2, q3, q4, q5, q6]); dist[comp] += 1
    status = "Correct" if comp == 6 and not r1 else ("Review-exec" if r1 else "Needs-fill")
    statc[status] += 1
    qa = ["Y" if q1 else "gap","Y" if q2 else "gap","Y" if q3 else "gap","Y" if q4 else "gap","Y" if q5 else "gap","Y" if q6 else "gap",
          f"{comp}/6", "; ".join(eh)[:200], status]
    rows.append(base + qa)
    audit_rows.append({"id": t.get("L5 Step ID"), "vs": vs, "division": l2, "l3": l3, "l4": l4,
                       "name": l5, "owner": owner, "participants": cap(parts, 200), "apps": cap(apps, 220),
                       "deliverable": cap(deliv, 160), "regs": cap(regs, 120), "exec_flag": bool(eh), "exec_roles": "; ".join(eh)[:160]})

# ---- coverage / reconciliation: synthetic vs bridge per division
syn_by_div = Counter(nk(r.get("L2 Division")) for r in v007)
brg_by_div = Counter(nk(r[1]) for r in rows)
brg_by_vs = Counter(nk(t.get("L2 Value Stream")) for t in bl5)

# ---- workbook
wb = openpyxl.Workbook()
YEL = PatternFill("solid", fgColor="FFF3B0"); HDR = PatternFill("solid", fgColor="1F3864"); HF = Font(color="FFFFFF", bold=True, size=11)
GAPF = PatternFill("solid", fgColor="F8D7DA"); REVF = PatternFill("solid", fgColor="FDE2C8"); OKF = PatternFill("solid", fgColor="E8F5E9")
ws = wb.active; ws.title = "MASTER + QA"
ws.append(COLS)
for j in range(1, len(COLS) + 1):
    c = ws.cell(1, j); c.fill = HDR; c.font = HF; c.alignment = Alignment(vertical="center", wrap_text=True)
ASSOC = list(range(8, 18))  # Owner..Automatability
for rec in rows:
    ws.append(rec); r = ws.max_row
    for j in ASSOC:
        if not ws.cell(r, j).value: ws.cell(r, j).fill = YEL
    st = rec[-1]; ws.cell(r, len(COLS)).fill = {"Correct": OKF, "Review-exec": REVF, "Needs-fill": GAPF}[st]
    for j in range(19, 25):  # Q cells
        if ws.cell(r, j).value == "gap": ws.cell(r, j).fill = GAPF
ws.freeze_panes = "F2"
widths = [13,22,24,26,32,36,5,22,32,14,28,20,12,30,24,12,11,13, 10,11,8,11,10,9,8,22,12]
for i, w in enumerate(widths, 1): ws.column_dimensions[get_column_letter(i)].width = w
ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{len(rows)+1}"

sc = wb.create_sheet("Scorecard")
sc.append(["SIX-QUESTIONS SCORECARD"]); sc.cell(1,1).font = Font(bold=True, size=13, color="1F3864")
sc.append([]); sc.append(["Total tasks", len(rows)])
sc.append(["R1 exec-on-task flags (review)", r1count])
sc.append([]); sc.append(["Completeness", "tasks"])
for k in range(6, -1, -1): sc.append([f"{k}/6", dist.get(k, 0)])
sc.append([]); sc.append(["QA status", "tasks"])
for k, v in statc.most_common(): sc.append([k, v])
sc.append([]); sc.append(["Column", "filled", "gap"])
for i, c in enumerate(BASECOLS[7:17], 7):
    f = sum(1 for r in rows if r[i]); sc.append([c, f, len(rows) - f])
for i, w in enumerate([34, 10, 10], 1): sc.column_dimensions[get_column_letter(i)].width = w
for rr in range(1, sc.max_row+1):
    if sc.cell(rr,1).value in ("Completeness","QA status","Column"): sc.cell(rr,1).font = Font(bold=True)

cv = wb.create_sheet("Coverage 710 vs 8355")
cv.append(["Answers the 710-vs-8355 question: vetted Bridge tasks vs synthetic L5, per value-stream division"])
cv.cell(1,1).font = Font(bold=True, color="1F3864")
cv.append([]); cv.append(["L2 Division", "Synthetic L5 (8,355)", "Bridge tasks (710)", "Bridge ÷ Synthetic", "Read"])
for dnk in sorted(div_display, key=lambda d: div_display[d]):
    s = syn_by_div.get(dnk, 0); b = brg_by_div.get(dnk, 0)
    ratio = round(b / s, 3) if s else 0
    read = "Bridge THIN here - expand" if (s and b / s < 0.04) else ("balanced" if s and b/s < 0.2 else "ok")
    cv.append([div_display[dnk], s, b, ratio, read])
cv.append([]); cv.append(["TOTAL", sum(syn_by_div.values()), sum(brg_by_div.values()), "", ""])
cv.append([]); cv.append(["Granular: Bridge tasks per Bridge value stream (29)"])
for vs in sorted(brg_by_vs, key=lambda v: -brg_by_vs[v]):
    cv.append([vs, "", brg_by_vs[vs]])
for i, w in enumerate([34, 18, 16, 16, 28], 1): cv.column_dimensions[get_column_letter(i)].width = w

rd = wb.create_sheet("README")
for ln in [
 ["MASTER + QA — six-questions correctness harness"],
 [""],
 ["WHAT THIS ADDS", "Per-task Q1-Q6 pass/gap, completeness /6, R1 exec-on-task flags, and a QA status. Plus a Scorecard and a 710-vs-8355 Coverage tab."],
 ["QA STATUS", "Correct = 6/6 and no exec flag (green). Review-exec = an Executive/Leadership role is on the task (orange) - confirm or replace. Needs-fill = missing answers (red)."],
 ["KNOWN GAP COLUMNS", "Applications/SoR (Q4), Checklist + Regulations (Q5), Automatability - blank+yellow on the Bridge base; these are the gap-fill targets."],
 ["R1 detection", "A role counts as exec if the role catalog marks it Executive/Leadership, or its name starts with Chief/Head of or contains Officer."],
 ["NEXT", "1) grounded gap-fill of Q4/Q5 from the app + regulation catalogs (verified). 2) expand THIN divisions flagged on the Coverage tab. 3) re-run this harness and watch the scorecard climb."],
] : rd.append(ln)
rd.column_dimensions["A"].width = 26; rd.column_dimensions["B"].width = 120
rd.cell(1,1).font = Font(bold=True, size=14, color="1F3864")

# audit bundle grouped by Bridge value stream (for the verified audit workflow)
bundle = defaultdict(list)
for a in audit_rows: bundle[a["vs"]].append(a)
json.dump(bundle, open(f"{O}/bundle_audit.json", "w", encoding="utf-8"), ensure_ascii=False)
print("bundle_audit.json:", len(bundle), "value streams,", len(audit_rows), "tasks")

out = "documents/value-streams/ABC-Insurance-Operating-Model-MASTER-v5-QA.xlsx"
wb.save(out)
print("SAVED", out)
print("tasks", len(rows), "| completeness dist", dict(sorted(dist.items())), "| status", dict(statc), "| R1 exec flags", r1count)
print("THIN divisions:", [div_display[d] for d in div_display if syn_by_div.get(d) and brg_by_div.get(d,0)/syn_by_div[d] < 0.04])
