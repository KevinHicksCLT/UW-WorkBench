# -*- coding: utf-8 -*-
# Merge re-fixed flagged owners into v11 -> v12 final, guardrail re-applied (universal role pool).
import json, re, sys
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from collections import Counter
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
O = "documents/value-streams/_master_build"
def nk(s): return re.sub(r"[^a-z0-9]+", " ", (s or "").lower()).strip()
def toks(s): return {w for w in re.findall(r"[a-z0-9]+", (s or "").lower()) if len(w) > 2}
JUNIOR = re.compile(r"associate|technician|assistant|trainee|junior|\bclerk\b|coordinator|representative|\brep\b|intern|processor|operator|\bsteward\b|data entry")
LVL = {0: "Junior IC", 1: "Specialist", 2: "Manager", 3: "Senior/Exec"}
def rank(n):
    n = (n or "").lower()
    if re.search(r"\bchief\b|head of |\bdirector\b|\bvp\b|president|general counsel|appointed actuary", n): return 3
    if re.search(r"\bmanager\b|\bhead\b|principal|\blead\b|superintendent", n): return 2
    if JUNIOR.search(n): return 0
    return 1
LEAD = re.compile(r"\bdepartment\b|enterprise|operating budget|\bbudget\b|strateg|operating model|governance|set (the )?standards?|\bcharter\b|policy framework|capital allocation|\broadmap\b|appetite framework|target operating|annual (operating )?plan")
APPROVE = re.compile(r"approve|sign.?off|authoriz|certif|attest|final (decision|approval)|go.?no.?go|board (approval|review)")
GOVERN = re.compile(r"operating plan|business plan|\bplan\b|planning|resource|staffing|capacity|headcount|oversight|prioriti|coordinate the program|workforce")
def altitude(name):
    n = name.lower()
    if APPROVE.search(n): return "approve"
    if LEAD.search(n): return "lead"
    if GOVERN.search(n): return "govern"
    return "execute"
allroles = json.load(open(f"{O}/all_roles.json", encoding="utf-8"))
def guardrail(name, l4, owner, pcs):
    alt = altitude(name); ro = rank(owner); mn = 2 if alt in ("lead", "approve") else 1
    if alt == "execute" and ro == 3:
        c = [p for p in pcs if 1 <= rank(p) <= 2]
        if c: return c[0], [owner] + [p for p in pcs if p != c[0]]
    if ro >= mn: return owner, pcs
    c = [p for p in pcs if rank(p) >= mn]
    if c: return c[0], [owner] + [p for p in pcs if p != c[0]]
    tt = toks(name + " " + l4); best = None; bs = -1
    for rr in allroles:
        if rank(rr["role"]) >= mn:
            ov = len(tt & toks(rr["role"] + " " + " ".join(rr.get("does", []))))
            if ov > bs: bs = ov; best = rr["role"]
    if best: return best, [owner] + [p for p in pcs if p != best]
    return owner, pcs

# re-fixed flagged owners
res = json.load(open(r"C:\Users\xando\AppData\Local\Temp\claude\C--Users-xando-Code-transform-platform\e7779fde-4039-436b-94a9-de90aeff60c5\tasks\w0cm3f993.output", encoding="utf-8"))["result"]["result"]
fixmap = {}
for grp in res:
    g = grp["group"]
    for t in grp.get("tasks", []):
        fixmap[(g, nk(t["l3"]), nk(t["l4"]), nk(t["name"]))] = t

src = "documents/value-streams/ABC-Insurance-Operating-Model-MASTER-v11-FINAL.xlsx"
ws0 = openpyxl.load_workbook(src)["MASTER - L5"]
H = [c.value for c in ws0[1]]; ix = {h: i for i, h in enumerate(H)}
COLS = H[:]  # same columns
rows = []; refixed = 0; stillflag = 0; jr = 0
for r in ws0.iter_rows(min_row=2, values_only=True):
    r = list(r)
    l2, l3, l4, name = r[ix["L2 Division"]], r[ix["L3 Process"]], r[ix["L4 Sub-Process"]], r[ix["L5 Task"]]
    fx = fixmap.get((l2, nk(l3), nk(l4), nk(name)))
    if fx:
        owner = fx["owner"]; pcs = fx.get("participants", [])
        owner, pcs = guardrail(name, l4, owner, pcs)
        r[ix["Owner / Lead Role"]] = owner
        r[ix["Roles (Participants)"]] = "; ".join(pcs[:3])
        r[ix["Owner Level"]] = LVL[rank(owner)]
        r[ix["Verified"]] = "OK" if fx.get("ok", True) else "FLAG"
        r[ix["Issues"]] = "" if fx.get("ok", True) else r[ix["Issues"]]
        refixed += 1
        if not fx.get("ok", True): stillflag += 1
    if rank(r[ix["Owner / Lead Role"]]) == 0: jr += 1
    rows.append(r)

wb = openpyxl.Workbook(); ws = wb.active; ws.title = "MASTER - L5"
HDR = PatternFill("solid", fgColor="1F3864"); HF = Font(color="FFFFFF", bold=True)
FLG = PatternFill("solid", fgColor="FDE2C8"); YEL = PatternFill("solid", fgColor="FFF3B0")
ws.append(COLS)
for j in range(1, len(COLS) + 1):
    c = ws.cell(1, j); c.fill = HDR; c.font = HF; c.alignment = Alignment(vertical="center", wrap_text=True)
for r in rows:
    ws.append(r); rr = ws.max_row
    if r[ix["Verified"]] == "FLAG": ws.cell(rr, ix["Verified"] + 1).fill = FLG
    if r[ix["Owner Level"]] == "Junior IC": ws.cell(rr, ix["Owner Level"] + 1).fill = YEL
ws.freeze_panes = "E2"
for i, w in enumerate([13,22,26,28,40,5,24,30,22,28,34,48,12,12,8,40], 1):
    if i <= len(COLS): ws.column_dimensions[get_column_letter(i)].width = w
ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{len(rows)+1}"
sc = wb.create_sheet("Scorecard"); lvlc = Counter(r[ix["Owner Level"]] for r in rows); vc = Counter(r[ix["Verified"]] for r in rows)
for ln in [["FINAL MASTER v12 — all flagged rows re-fixed"], [],
           ["L5 tasks", len(rows)], ["Flagged rows re-fixed", refixed], ["Still flagged (review)", stillflag],
           ["Junior-as-lead", jr], [], ["Owner level", ""]] + [[k, lvlc.get(k, 0)] for k in ["Senior/Exec","Manager","Specialist","Junior IC"]] + [[], ["Verified", ""]] + [[k, vc.get(k, 0)] for k in vc]:
    sc.append(ln)
sc.cell(1,1).font = Font(bold=True, size=13, color="1F3864")
for i, w in enumerate([30, 12], 1): sc.column_dimensions[get_column_letter(i)].width = w
out = "documents/value-streams/ABC-Insurance-Operating-Model-MASTER-v12-FINAL.xlsx"
wb.save(out)
print("SAVED", out)
print("tasks", len(rows), "| re-fixed", refixed, "| still flagged", stillflag, "| junior-as-lead", jr)
print("owner levels:", dict(lvlc), "| verified:", dict(vc))
def show(sub):
    for r in rows:
        if sub.lower() in (r[ix["L5 Task"]] or "").lower():
            print(f"  [{r[ix['L2 Division']]}] {r[ix['L5 Task']]} -> {r[ix['Owner / Lead Role']]} ({r[ix['Owner Level']]})"); return
for s in ["Translate pricing signals into proposed appetite","Post booked reserve journal","Track actuarial continuing"]: show(s)
