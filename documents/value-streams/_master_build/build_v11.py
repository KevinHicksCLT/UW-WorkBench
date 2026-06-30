# -*- coding: utf-8 -*-
# Final merge: precise LLM owners + deterministic SENIORITY GUARDRAIL (promote true-juniors to a qualified
# participant; keep specialists/managers/seniors). Analyst = specialist (owns analysis). Keep v10 apps.
import json, re, sys
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from collections import Counter
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
def nk(s): return re.sub(r"[^a-z0-9]+", " ", (s or "").lower()).strip()
# TRUE junior support roles (Analyst removed — analysts own analytical work)
JUNIOR = re.compile(r"associate|technician|assistant|trainee|junior|\bclerk\b|coordinator|representative|\brep\b|intern|processor|operator|\bsteward\b|data entry")
LVL = {0: "Junior IC", 1: "Specialist", 2: "Manager", 3: "Senior/Exec"}
def rank(n):
    n = (n or "").lower()
    if re.search(r"\bchief\b|head of |\bdirector\b|\bvp\b|president|general counsel|appointed actuary", n): return 3
    if re.search(r"\bmanager\b|\bhead\b|principal|\blead\b|superintendent", n): return 2
    if JUNIOR.search(n): return 0
    return 1
LEAD = re.compile(r"\bdepartment\b|enterprise|operating budget|\bbudget\b|strateg|operating model|governance|set (the )?standards?|\bcharter\b|policy framework|capital allocation|\broadmap\b|appetite framework|target operating|annual (operating )?plan")
APPROVE = re.compile(r"approve|sign.?off|authoriz|certif|attest|final (decision|approval)|go.?no.?go|board (approval|review)")
GOVERN = re.compile(r"operating plan|business plan|\bplan\b|planning|resource|staffing|capacity|headcount|oversight|prioriti|coordinate the program|workforce")
def altitude(name):
    n = name.lower()
    if APPROVE.search(n): return "approve"
    if LEAD.search(n): return "lead"
    if GOVERN.search(n): return "govern"
    return "execute"
rref = json.load(open("documents/value-streams/_master_build/role_ref_by_division.json", encoding="utf-8"))
def toks(s):
    return {w for w in re.findall(r"[a-z0-9]+", (s or "").lower()) if len(w) > 2}
def guardrail(name, l4, owner, pcs, divroles):
    """Ensure owner meets the seniority band for the task; promote via participant, else via division pool."""
    alt = altitude(name); ro = rank(owner)
    mn = 2 if alt in ("lead", "approve") else 1   # minimum acceptable rank
    # demote an exec sitting on routine execution
    if alt == "execute" and ro == 3:
        cand = [p for p in pcs if 1 <= rank(p) <= 2]
        if cand:
            new = cand[0]; return new, [owner] + [p for p in pcs if p != new]
    if ro >= mn:
        return owner, pcs
    # promote: prefer a qualifying participant
    cand = [p for p in pcs if rank(p) >= mn]
    if cand:
        new = cand[0]; return new, [owner] + [p for p in pcs if p != new]
    # fallback: best-matching non-junior from the division pool
    tt = toks(name + " " + l4); best = None; bs = -1
    for rr2 in divroles:
        if rank(rr2["role"]) >= mn:
            ov = len(tt & toks(rr2["role"] + " " + " ".join(rr2.get("does", []))))
            if ov > bs: bs = ov; best = rr2["role"]
    if best:
        return best, [owner] + [p for p in pcs if p != best]
    return owner, pcs

F = r"C:\Users\xando\AppData\Local\Temp\claude\C--Users-xando-Code-transform-platform\e7779fde-4039-436b-94a9-de90aeff60c5\tasks\w0rtxder3.output"
res = json.load(open(F, encoding="utf-8"))["result"]
omap = {}
for g in res["groups"]:
    for l4g in g.get("l4groups", []):
        for t in l4g.get("tasks", []):
            omap[(g["l3key"], l4g["l4"], nk(t["name"]))] = t

src = "documents/value-streams/ABC-Insurance-Operating-Model-MASTER-v10-FIXED.xlsx"
ws0 = openpyxl.load_workbook(src)["MASTER - L5 (fixed)"]
H = [c.value for c in ws0[1]]; ix = {h: i for i, h in enumerate(H)}
COLS = ["L1 Segment","L2 Division","L3 Process","L4 Sub-Process","L5 Task","isTask","Owner / Lead Role","Roles (Participants)",
        "Primary App / SoR","Supporting Apps","Deliverable","Checklist","Automatability","Owner Level","Verified","Issues"]
rows = []; applied = 0; flagged = 0; jr = 0; promoted = 0
for r in ws0.iter_rows(min_row=2, values_only=True):
    l1,l2,l3,l4,name = r[ix["L1 Segment"]],r[ix["L2 Division"]],r[ix["L3 Process"]],r[ix["L4 Sub-Process"]],r[ix["L5 Task"]]
    owner = r[ix["Owner / Lead Role"]]; pcs = [p.strip() for p in (r[ix["Roles (Participants)"]] or "").split(";") if p.strip()]
    verified = "n/a"; issues = ""
    pe = omap.get((f"{l2} || {l3}", l4, nk(name)))
    if pe:
        owner = pe["owner"]; pcs = pe.get("participants", []); applied += 1
        verified = "OK" if pe.get("ok") else "FLAG"
        if not pe.get("ok"): flagged += 1; issues = pe.get("issues", "")
    o2, pcs = guardrail(name, l4, owner, pcs, rref.get(l2, []))
    if o2 != owner: promoted += 1
    owner = o2
    if rank(owner) == 0: jr += 1
    rows.append([l1,l2,l3,l4,name,"Y",owner,"; ".join(pcs[:3]),r[ix["Primary App / SoR"]],r[ix["Supporting Apps"]],
                 r[ix["Deliverable"]],r[ix["Checklist"]],r[ix["Automatability"]],LVL[rank(owner)],verified,issues])

wb = openpyxl.Workbook(); ws = wb.active; ws.title = "MASTER - L5"
HDR = PatternFill("solid", fgColor="1F3864"); HF = Font(color="FFFFFF", bold=True)
FLG = PatternFill("solid", fgColor="FDE2C8"); YEL = PatternFill("solid", fgColor="FFF3B0")
ws.append(COLS)
for j in range(1, len(COLS)+1):
    c=ws.cell(1,j); c.fill=HDR; c.font=HF; c.alignment=Alignment(vertical="center",wrap_text=True)
for r in rows:
    ws.append(r); rr=ws.max_row
    if r[14]=="FLAG": ws.cell(rr,15).fill=FLG
    if r[13]=="Junior IC": ws.cell(rr,14).fill=YEL
ws.freeze_panes="E2"
for i,w in enumerate([13,22,26,28,40,5,24,30,22,28,34,48,12,12,8,40],1): ws.column_dimensions[get_column_letter(i)].width=w
ws.auto_filter.ref=f"A1:{get_column_letter(len(COLS))}{len(rows)+1}"
sc=wb.create_sheet("Scorecard"); lvlc=Counter(r[13] for r in rows)
for ln in [["FINAL MASTER — owners precise + seniority-guarded + apps corrected"],[],
           ["L5 tasks",len(rows)],["Precise-owner applied",applied],["Owners promoted off junior",promoted],
           ["Verifier-flagged (review)",flagged],["Junior-as-lead remaining",jr],[],["Owner level",""]] + [[k,lvlc.get(k,0)] for k in ["Senior/Exec","Manager","Specialist","Junior IC"]]:
    sc.append(ln)
sc.cell(1,1).font=Font(bold=True,size=13,color="1F3864")
for i,w in enumerate([34,12],1): sc.column_dimensions[get_column_letter(i)].width=w
out="documents/value-streams/ABC-Insurance-Operating-Model-MASTER-v11-FINAL.xlsx"
wb.save(out)
print("SAVED",out)
print("tasks",len(rows),"| applied",applied,"| promoted off junior",promoted,"| flagged",flagged,"| junior-as-lead",jr)
print("owner levels:",dict(lvlc))
def show(sub):
    for r in rows:
        if sub.lower() in r[4].lower(): print(f"  [{r[1]}] {r[4]} -> {r[6]} ({r[13]}) | {r[8]}"); return
for s in ["department operating budget","Extract in-force and claims data","go/no-go review and record launch","Capture structured loss facts","Calculate indicated rate change"]: show(s)
# any remaining junior owners?
print("\nremaining junior-as-lead samples:")
n=0
for r in rows:
    if r[13]=="Junior IC":
        print(f"  [{r[1]}] {r[4]} -> {r[6]}"); n+=1
        if n>=6: break
