# -*- coding: utf-8 -*-
import json, re, sys
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
O = "documents/value-streams/_master_build"
def L(n): return json.load(open(f"{O}/{n}", encoding="utf-8"))
def nkey(s): return re.sub(r"[^a-z0-9]+", " ", (s or "").lower()).strip()

v007 = L("wb_v007.json"); v7 = L("wb_v7.json")
bl4 = L("wb_bridge_l4.json"); bl5 = L("wb_bridge_l5.json")
stds = L("wb_standards.json"); bext = L("wb_ext.json")
PRODF = "C:/Users/xando/.claude/projects/C--Users-xando-Code-transform-platform/ef5e9499-acfa-4c2b-a6ad-9f6711879ab5/tool-results/mcp-Neon-run_sql-1782174530544.txt"
prod = json.load(open(PRODF, encoding="utf-8"))[0]["data"]
vsreg = L("db_prod_vsreg.json")
db_role = L("db_role.json"); db_app = L("db_app.json"); db_deliv = L("db_deliv.json")
db_standard = L("db_standard.json")

# production VS-grain aggregates (roles/apps/metrics/reg) grouped by value stream
def split(s): return [x.strip() for x in (s or "").split(";") if x.strip()]
vsagg = {}
for r in prod:
    vs = r.get("vs")
    if not vs: continue
    a = vsagg.setdefault(vs, {"roles": set(), "apps": set(), "metrics": set(), "reg_n": 0})
    a["roles"].update(split(r.get("vs_roles"))); a["apps"].update(split(r.get("vs_apps")))
    a["metrics"].update(split(r.get("vs_metrics")))
    try: a["reg_n"] = max(a["reg_n"], int(r.get("reg_n") or 0))
    except: pass
vsreg_map = {nkey(r["vs"]): r for r in vsreg}

# L2 Division (locked spine) -> production Value Stream(s) crosswalk
XWALK = {
 "actuarial": ["Actuarial Pricing, Reserving & Capital Modeling"],
 "business operations": ["Policy Administration & Servicing", "Service Operations, Incident & Production Support"],
 "call center": ["Customer Service, Complaints & Experience"],
 "claims": ["Claims Intake-to-Settlement", "Claims Recoveries & Subrogation", "Life & Annuity Claims"],
 "corporate functions operations": ["Enterprise Strategy & Portfolio Management", "Third-Party & Vendor Management"],
 "cybersecurity iam": ["Cybersecurity, Identity & Resilience"],
 "data ai": ["Data, Analytics & AI Management", "MLOps / ML Lifecycle Management", "AIOps & Intelligent Operations"],
 "finance investments": ["Finance, Treasury & Capital Management", "Investment & Asset Management", "FinOps / Cloud Financial Management"],
 "human resources talent": ["Talent & Workforce Management"],
 "legal corporate governance": ["Legal, Governance & Privacy Management"],
 "product delivery": ["Product & Proposition Management", "Technology Delivery & Change"],
 "program management office": ["Change Management & Adoption", "Enterprise Strategy & Portfolio Management"],
 "reinsurance": ["Reinsurance & Retrocession Management"],
 "risk compliance audit": ["Risk, Compliance & Regulatory Management", "Audit & Assurance"],
 "sales distribution marketing": ["Distribution & Channel Management", "Marketing, Growth & Customer Insights"],
 "technology engineering": ["Technology Strategy, Architecture & Delivery", "Technology Delivery & Change", "Service Operations, Incident & Production Support", "FinOps / Cloud Financial Management"],
 "underwriting": ["Submission-to-Bind / Underwriting", "Delegated Authority Management"],
}
v7i = {}
for r in v7:
    v7i[(nkey(r.get("L1 Segment")), nkey(r.get("L2 Division")), nkey(r.get("L3 Process")), nkey(r.get("L4 Sub-Process")), nkey(r.get("L5 Step")))] = r
std_map = {}
for r in stds:
    d = r.get("Department / Team")
    if d: std_map[nkey(d)] = {"count": r.get("Standards Count"), "owner": r.get("Primary Owner")}
# curated L2 Division -> standards-domain keyword (avoids loose token misfires)
STD_X = {
 "actuarial": "actuarial", "claims": "claims", "underwriting": "underwriting",
 "technology engineering": "engineering", "data ai": "data", "finance investments": "finance",
 "human resources talent": "hr", "risk compliance audit": "compliance", "legal corporate governance": "legal",
 "cybersecurity iam": "cyber", "product delivery": "pmo", "program management office": "pmo",
 "business operations": "operations", "call center": "operations", "corporate functions operations": "operations",
}

def ext_match(l2, l3):
    hits = []
    for r in bext:
        rv = nkey(r.get("Related Value Stream")) + " " + nkey(r.get("Division / Function"))
        if (nkey(l2) and nkey(l2) in rv) or (nkey(l3) and nkey(l3) in rv):
            hits.append(f"{r.get('External Party Type')}/{r.get('External Role')} ({r.get('Interaction Type')})")
    return "; ".join(dict.fromkeys(hits))[:300]
def first(*xs):
    for x in xs:
        if x and str(x).strip(): return str(x).strip()
    return ""
def cap(s, n=600): return s if len(s) <= n else s[:n] + " ..."

COLS = ["L1 Segment","L2 Division","L3 Process","L3#","L4 Sub-Process","L4#","L5 Step","L5#","isTask",
        "DB Value Stream (mapped)","Roles (Participants)","Owner / Lead Role","Applications / SoR","Deliverables",
        "Inputs","Outputs","Standards","Regulations","Metrics","External Interactions",
        "Checklist","How / Test Pattern","Automatability","Match Level","Sources","DB-backed","Assoc Status"]
master = []; stat = {"Associated": 0, "Partial": 0, "GAP": 0}; colfill = {c: 0 for c in COLS}
YELLOW_COLS = list(range(10, 24))  # 1-based: DB VS .. Automatability
for row in v007:
    l1 = row.get("L1 Segment"); l2 = row.get("L2 Division"); l3 = row.get("L3 Process"); l4 = row.get("L4 Sub-Process"); l5 = row.get("L5 Step")
    v7r = v7i.get((nkey(l1), nkey(l2), nkey(l3), nkey(l4), nkey(l5)))
    mapped = [vs for vs in XWALK.get(nkey(l2), []) if vs in vsagg]
    src = set(); db_backed = "N"
    db_roles = set(); db_apps = set(); db_metrics = set(); reg_n = 0
    for vs in mapped:
        a = vsagg[vs]; db_roles |= a["roles"]; db_apps |= a["apps"]; db_metrics |= a["metrics"]; reg_n = max(reg_n, a["reg_n"])
    dbvs = "; ".join(mapped)
    if mapped: db_backed = "Y"; src.add("DB")
    roles = first("; ".join(sorted(db_roles)), (v7r or {}).get("Roles"))
    if roles and not db_roles: src.add("v7")
    owner = ""  # no source maps to the locked spine at owner grain
    apps = first("; ".join(sorted(db_apps)), (v7r or {}).get("Application / SOR Type Examples"))
    if apps and not db_apps: src.add("v7")
    deliv = first((v7r or {}).get("Deliverables"))
    if deliv: src.add("v7")
    inputs = ""; outputs = ""  # step-grain DB/Bridge could not map (taxonomy divergence)
    std = ""
    kw = STD_X.get(nkey(l2))
    if kw:
        for dk, dv in std_map.items():
            if kw in dk:
                std = f"{dv['count']} standards - {dv['owner']}"; src.add("Std"); break
    regs = ""
    if reg_n > 0:
        for vs in mapped:
            rt = vsreg_map.get(nkey(vs), {}).get("reg_titles", "")
            if rt: regs = f"{reg_n} requirements - {rt}"; break
        if not regs: regs = f"{reg_n} requirements"
        src.add("DB")
    metrics = "; ".join(sorted(db_metrics))
    ext = ext_match(l2, l3)
    if ext: src.add("Bridge")
    checklist = first((v7r or {}).get("Checklist"), row.get("Checklist"))
    if checklist: src.add("v7")
    how = first((v7r or {}).get("How"), row.get("How"), (v7r or {}).get("Test by Pattern"))
    ml = "VS-crosswalk" if db_backed == "Y" else ("v7" if v7r else "none")
    rec = [l1, l2, l3, row.get("L3#"), l4, row.get("L4#"), l5, row.get("L5#"), "Y",
           cap(dbvs, 80), cap(roles), owner, cap(apps), cap(deliv), inputs, outputs, std, cap(regs), cap(metrics), cap(ext, 300),
           cap(checklist, 800), cap(how, 400), "", ml, "; ".join(sorted(src)) or "GAP", db_backed, ""]
    nfill = sum(1 for i in [10, 12, 13, 16, 17, 18, 19] if rec[i])  # roles,apps,deliv,std,reg,metric,ext
    rec[26] = "Associated" if nfill >= 4 else ("Partial" if nfill >= 1 else "GAP")
    stat[rec[26]] += 1
    for i, c in enumerate(COLS):
        if rec[i]: colfill[c] += 1
    master.append(rec)

wb = openpyxl.Workbook()
YEL = PatternFill("solid", fgColor="FFF3B0"); HDR = PatternFill("solid", fgColor="1F3864"); HF = Font(color="FFFFFF", bold=True, size=11)
GAPF = PatternFill("solid", fgColor="F8D7DA"); PARTF = PatternFill("solid", fgColor="FFF8E1"); OKF = PatternFill("solid", fgColor="E8F5E9")
DBOK = PatternFill("solid", fgColor="D5E8FF")
ws = wb.active; ws.title = "MASTER - L5"; ws.append(COLS)
for j in range(1, len(COLS) + 1):
    c = ws.cell(1, j); c.fill = HDR; c.font = HF; c.alignment = Alignment(vertical="center", wrap_text=True)
for rec in master:
    ws.append(rec); r = ws.max_row
    for j in YELLOW_COLS:
        if not ws.cell(r, j).value: ws.cell(r, j).fill = YEL
    ws.cell(r, 26).fill = DBOK if rec[25] == "Y" else YEL
    ws.cell(r, 27).fill = {"GAP": GAPF, "Partial": PARTF, "Associated": OKF}[rec[26]]
ws.freeze_panes = "J2"
for i, w in enumerate([13,18,24,5,24,5,28,5,6,26,34,16,30,30,18,18,22,34,30,26,34,26,9,13,16,9,11], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{len(master)+1}"

rd = wb.create_sheet("README & Legend")
db_y = sum(1 for r in master if r[25] == "Y")
lines = [
 ["ABC Insurance - Operating-Model MASTER (v3 data-model staging sheet)", ""],
 ["", ""],
 ["PURPOSE", "One reviewable sheet to QA every value-stream to L5 node and its associations BEFORE loading the DB."],
 ["LOCKED", "L1 Segment > L2 Division > L3 Process > L4 Sub-Process > L5 Step are LOCKED (not edited here)."],
 ["GRAIN", f"One row per L5 step (Task). {len(master)} rows."],
 ["", ""],
 ["*** KEY DATA-QUALITY FINDING ***", "The locked v007 spine was RE-DECOMPOSED: its L3/L4 names match the production DB 0 of 867 and the Bridge workbook 0 of 867. So step-level DB/Bridge associations CANNOT be auto-mapped by name."],
 ["What this means", "Production data was reassociated at the VALUE-STREAM grain via an L2 Division -> production VS crosswalk (col J). Step-grain precision (Owner, Inputs, Outputs, step-exact apps/deliverables) could NOT be carried and is YELLOW = needs rebuild."],
 ["DB-backed rows", f"{db_y} of {len(master)} L5 rows received production DB associations via the VS crosswalk."],
 ["", ""],
 ["COLOR KEY", ""],
 ["  Yellow cell", "NOT ASSOCIATED - needs data or review (the gaps you asked to flag)."],
 ["  Blue cell (DB-backed=Y)", "Production DB reassociated this node (at VS grain)."],
 ["  Status Associated (green) / Partial (amber) / GAP (red)", "4+ / 1-3 / 0 key associations."],
 ["", ""],
 ["SOURCES column", ""],
 ["  DB", "Production Neon branch (br-curly-poetry), VS-grain: RoleValueStream / ApplicationValueStream / Metric / RequirementValueStream via L2->VS crosswalk."],
 ["  v7", "ABC-Insurance-Operating-Model-v7_062226.xlsx populated columns (same locked spine) - Roles, Deliverables, Apps/SOR, Checklist, How."],
 ["  Std", "Bridge Standards Index (by department/domain)."],
 ["  Bridge", "AI Transformation Bridge Input.xlsx External Interactions (by VS keyword)."],
 ["  GAP", "No source carried an association for this node."],
 ["", ""],
 ["NOTE working branch", "Active app DB branch (proud-lab/ep-raspy-boat) has the 8,355-step spine but its connection tables are WIPED (0 rows). Associations were re-read from PRODUCTION (br-curly-poetry)."],
 ["NOT INCLUDED YET", "Owner/Lead role, Inputs, Outputs, step-exact apps/deliverables (taxonomy-divergence gaps); Automatability scoring; TestingTemplate; AppChargeback; KnowledgeBase/AgentSkill; per-instance display labels."],
 ["NEXT", "Review yellow cells; confirm/repair the L2->VS crosswalk (col J); decide whether to re-key the locked spine to DB codes so step-level DB associations can map."],
]
for ln in lines: rd.append(ln)
rd.column_dimensions["A"].width = 40; rd.column_dimensions["B"].width = 120
rd.cell(1, 1).font = Font(bold=True, size=14, color="1F3864")
rd.cell(7, 1).font = Font(bold=True, color="C0392B"); rd.cell(7, 2).font = Font(bold=True, color="C0392B")
rd.cell(12, 1).fill = YEL; rd.cell(13, 1).fill = DBOK

gs = wb.create_sheet("Gaps Summary")
gs.append(["Overall", "L5 nodes", len(master)])
for k in ("Associated", "Partial", "GAP"): gs.append(["", k, stat[k]])
gs.append(["", "DB-backed (VS crosswalk)", db_y])
gs.append([]); gs.append(["Column fill rate", "Filled", "of " + str(len(master))])
for c in COLS[9:23]: gs.append(["", c, colfill[c]])
gs.append([]); gs.append(["By L2 Division", "L5 nodes", "Associated", "Partial", "GAP", "DB-backed"])
from collections import defaultdict
byd = defaultdict(lambda: [0, 0, 0, 0, 0])
for rec in master:
    d = rec[1] or "(blank)"; byd[d][0] += 1; byd[d][{"Associated": 1, "Partial": 2, "GAP": 3}[rec[26]]] += 1
    if rec[25] == "Y": byd[d][4] += 1
for d in sorted(byd):
    v = byd[d]; gs.append([d, v[0], v[1], v[2], v[3], v[4]])
for i, w in enumerate([30, 12, 12, 12, 10, 12], 1): gs.column_dimensions[get_column_letter(i)].width = w
gs.cell(1, 1).font = Font(bold=True)

def cat(name, rows, cols, keys):
    s = wb.create_sheet(name); s.append(cols)
    for j in range(1, len(cols) + 1): s.cell(1, j).fill = HDR; s.cell(1, j).font = HF
    for r in rows: s.append([r.get(k, "") for k in keys])
    s.freeze_panes = "A2"
    for j, w in enumerate([34, 18, 18, 18, 22, 40], 1):
        if j <= len(cols): s.column_dimensions[get_column_letter(j)].width = w
    s.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"
cat("Catalog - Roles", db_role, ["Role", "Family", "Level", "Domain", "Primary VS", "MgrId"], ["name", "fam", "lvl", "dom", "pvs", "mgr"])
cat("Catalog - Applications", db_app, ["Application", "Kind", "Category", "Vendor", "SoR", "Code"], ["name", "kind", "category", "vendor", "sor", "code"])
cat("Catalog - Deliverables", db_deliv, ["Deliverable", "Type", "Owner", "Status", "Value Stream"], ["title", "type", "owner", "status", "vs"])
cat("Catalog - Standards", db_standard, ["Department", "Count", "Owner", "Mission", "Scope"], ["department", "count", "owner", "mission", "scope"])
cat("Catalog - Regulations", vsreg, ["Value Stream", "# Requirements", "Example Regimes"], ["vs", "reg_n", "reg_titles"])
cat("Catalog - External", bext, ["Party Type", "External Role", "Internal Owner", "Interaction", "Related VS", "Dependency"], ["External Party Type", "External Role", "Internal Role Owner", "Interaction Type", "Related Value Stream", "Dependency Type"])

out = "documents/value-streams/ABC-Insurance-Operating-Model-MASTER-v3.xlsx"
wb.save(out)
print("SAVED", out)
print("rows", len(master), "status", stat, "db_backed", db_y)
print("col fill:", {c: colfill[c] for c in COLS[9:23]})
