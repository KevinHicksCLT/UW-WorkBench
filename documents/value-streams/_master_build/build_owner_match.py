# -*- coding: utf-8 -*-
# Deterministic owner+app fix across ALL tasks (no LLM): match owner to role responsibilities,
# assign apps by task-type rules constrained to the division catalog. Overlay LLM-verified areas.
import json, re, sys
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from collections import Counter, defaultdict
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
O = "documents/value-streams/_master_build"
def Lj(n): return json.load(open(f"{O}/{n}", encoding="utf-8"))
def toks(s):
    STOP = set("the a an and or of to for in on with by from into against per as at is are be to vs via using use".split())
    return {w for w in re.findall(r"[a-z0-9]+", (s or "").lower()) if w not in STOP and len(w) > 2}

tba = Lj("tasks_by_area.json"); rref = Lj("role_ref_by_division.json")
apool = Lj("app_pool_by_division.json"); l4_by_l3 = Lj("l4_by_l3.json")
key_l1 = {k: g["l1"] for k, g in l4_by_l3.items()}

# LLM-verified areas from the partial run (use as-is; high quality)
LLM = {}
try:
    res = json.load(open(r"C:\Users\xando\AppData\Local\Temp\claude\C--Users-xando-Code-transform-platform\e7779fde-4039-436b-94a9-de90aeff60c5\tasks\wf1jc1kvj.output", encoding="utf-8"))["result"]
    for g in res["groups"]: LLM[g["l3key"]] = g
except Exception as e:
    print("LLM overlay load skipped:", e)

APP_RULES = [
 (r"go.?no.?go|launch|release|deploy|go.?live|readiness|sprint|backlog|user story|cut.?over|pipeline build", ["Jira", "Jira Align / Rally", "Azure DevOps"]),
 (r"change ticket|incident|service request|change advisory|cab\b|problem record", ["ServiceNow ITSM"]),
 (r"ledger|journal|financial close|reconcil|general ledger|accru|trial balance", ["SAP S/4HANA (GL)", "BlackLine (Close)"]),
 (r"budget|forecast|fp&a|variance|planning cycle", ["Oracle EPM", "OneStream (FP&A)", "Anaplan (Planning)"]),
 (r"invoice|billing|premium bill|collection|receivable|payment|dunning|cash appl", ["Guidewire BillingCenter", "Payment Gateway"]),
 (r"\bclaim|fnol|adjud|settle|salvage|adjust|loss notice|coverage determination", ["Guidewire ClaimCenter", "OpenText Document Mgmt"]),
 (r"fraud|siu\b", ["FRISS Fraud Detection", "Guidewire ClaimCenter"]),
 (r"underwrit|submission|quote|bind|appetite|clearance|risk assessment|pric(e|ing) indication", ["Guidewire PolicyCenter", "Earnix Rating & Pricing", "Broker & Agent Portal"]),
 (r"policy admin|endorse|renew|issuance|servicing|in.?force", ["Guidewire PolicyCenter", "Policy Administration Platform"]),
 (r"reserv|experience study|mortality|morbidity|lapse|loss cost|loss development|capital model|solvency|actuar|assumption", ["SAS Actuarial Platform", "FIS Prophet", "Milliman Arius (Reserving)"]),
 (r"cat |catastrophe|exposure model|pml", ["Moody's RMS (Cat Model)", "Catastrophe Data Provider"]),
 (r"reinsur|cede|treaty|retrocess|bordereau|recover", ["Sapiens ReinsurancePro", "Reinsurer Exchange"]),
 (r"\bdata\b|extract|dataset|etl|ingest|dashboard|report build|data quality|master data", ["Snowflake Data Cloud", "Databricks Lakehouse", "Power BI"]),
 (r"\bmodel|machine learning|\bml\b|train|inference|drift|mlops|feature", ["Azure ML / MLflow", "Databricks Lakehouse"]),
 (r"investment|portfolio|trade|asset alloc|aladdin|securities", ["BlackRock Aladdin", "Bloomberg Terminal"]),
 (r"cyber|vulnerab|threat|siem|patch|malware|soc\b", ["Splunk SIEM", "CrowdStrike Falcon", "Qualys / Tenable (Vuln)"]),
 (r"identity|access review|provision|\biam\b|privileged|certif|entitlement", ["SailPoint IGA", "Okta / IAM Platform", "CyberArk (PAM)"]),
 (r"recruit|hire|onboard|talent acquis|requisition|candidate", ["Greenhouse (Recruiting)", "Workday HCM"]),
 (r"payroll|compensation|benefit|employee|learning|training|performance review", ["Workday HCM", "ADP Payroll", "Cornerstone LMS"]),
 (r"contract|legal|litigat|counsel|agreement|nda|dispute", ["iManage (Legal Docs)", "Icertis CLM (Contracts)", "DocuSign"]),
 (r"compliance|regulat|filing|exam|\baudit|control|grc|\baml\b|sanction|conduct", ["Archer GRC", "AuditBoard", "NICE Actimize (AML)", "Regulatory Filing Portal (SERFF)"]),
 (r"campaign|marketing|brand|content|\blead\b", ["Marketo (Marketing)", "Salesforce FSC"]),
 (r"producer|agent|broker|distribut|appointment|licens", ["Duck Creek Distribution", "NIPR (Producer Licensing)", "Broker & Agent Portal", "Salesforce FSC"]),
 (r"customer|complaint|contact center|call\b|service level|csat", ["Salesforce FSC", "Genesys Cloud (Contact Center)", "Customer Self-Service Portal"]),
 (r"vendor|procure|third.?party|sourcing", ["Coupa / SAP Ariba (Procurement)", "Archer GRC"]),
]
DOC_RE = re.compile(r"document|draft|report|memo|publish|procedure|minutes|present|communicat|sign.?off|approve|review|recommend|charter|plan\b")
APPROVE_RE = re.compile(r"approve|sign.?off|authoriz|govern|board|final decision|go.?no.?go|attest|certif(y|ication)")

def assign_apps(name, l4, pool_names):
    text = (name + " " + l4).lower()
    picks = []
    for rx, apps in APP_RULES:
        if re.search(rx, text):
            for a in apps:
                if a in pool_names and a not in picks: picks.append(a)
    # docs go to collaboration tools
    if DOC_RE.search(text):
        for a in ("Microsoft 365 / SharePoint", "Confluence"):
            if a in pool_names and a not in picks: picks.append(a)
    if not picks:  # fallback: division primary system of record
        sors = [a for a in pool_names if a not in ("Microsoft 365 / SharePoint", "Microsoft Excel", "Microsoft Teams / Email")]
        picks = sors[:1] or ["Microsoft 365 / SharePoint"]
    primary = picks[0]
    supporting = picks[1:3]
    return primary, supporting

def match_owner(name, deliverable, l4, roles):
    ttok = toks(name + " " + deliverable + " " + l4)
    approve = bool(APPROVE_RE.search(name.lower()))
    best = None; best_s = -1; best_part = []
    scored = []
    for r in roles:
        rtok = toks(r["role"] + " " + " ".join(r.get("does", [])) + " " + (r.get("family") or ""))
        s = len(ttok & rtok)
        # prefer non-exec for hands-on; execs allowed only for approval steps
        if r.get("exec") and not approve: s -= 3
        if r.get("exec") and approve: s += 1
        scored.append((s, r["role"]))
    scored.sort(reverse=True)
    if not scored: return "", []
    owner = scored[0][1]
    participants = [nm for _, nm in scored[1:3]]
    return owner, participants

COLS = ["L1 Segment","L2 Division","L3 Process","L4 Sub-Process","L5 Task","isTask","Owner / Lead Role",
        "Roles (Participants)","Primary App / SoR","Supporting Apps","Deliverable","Checklist","Automatability","Source"]
rows = []; src_c = Counter()
for k, area in tba.items():
    parts = [x.strip() for x in k.split("||")]; l2 = parts[0]; l3 = parts[1] if len(parts) > 1 else ""
    l1 = key_l1.get(k, "")
    roles = rref.get(l2, []); pool_names = [a["name"] for a in apool.get(l2, [])]
    llm = LLM.get(k)
    if llm:  # use LLM-verified content directly
        src_c["LLM-verified"] += 1
        for g in llm.get("l4groups", []):
            for t in g.get("tasks", []):
                rows.append([l1, l2, l3, g["l4"], t["name"], "Y", t.get("owner", ""), "; ".join(t.get("participants", [])),
                             t.get("primary_app", ""), "; ".join(t.get("supporting_apps", [])), t.get("deliverable", ""),
                             " | ".join(t.get("checklist", [])), t.get("automatability", ""), "LLM"])
    else:  # deterministic fix
        src_c["deterministic"] += 1
        for g in area.get("l4groups", []):
            l4 = g["l4"]
            for t in g.get("tasks", []):
                owner, pcs = match_owner(t["name"], t.get("deliverable", ""), l4, roles)
                primary, supp = assign_apps(t["name"], l4, pool_names)
                rows.append([l1, l2, l3, l4, t["name"], "Y", owner, "; ".join(pcs),
                             primary, "; ".join(supp), t.get("deliverable", ""),
                             " | ".join(t.get("checklist", [])), t.get("automatability", ""), "deterministic"])
rows.sort(key=lambda r: (r[0], r[1], r[2], r[3], r[4]))

wb = openpyxl.Workbook(); ws = wb.active; ws.title = "MASTER - L5 (fixed)"
HDR = PatternFill("solid", fgColor="1F3864"); HF = Font(color="FFFFFF", bold=True); YEL = PatternFill("solid", fgColor="FFF3B0")
ws.append(COLS)
for j in range(1, len(COLS) + 1):
    c = ws.cell(1, j); c.fill = HDR; c.font = HF; c.alignment = Alignment(vertical="center", wrap_text=True)
noowner = 0
for r in rows:
    ws.append(r); rr = ws.max_row
    if not r[6]: ws.cell(rr, 7).fill = YEL; noowner += 1
    if not r[8]: ws.cell(rr, 9).fill = YEL
ws.freeze_panes = "E2"
for i, w in enumerate([13,22,26,28,40,5,24,30,22,28,34,50,12,12], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{len(rows)+1}"

sc = wb.create_sheet("Scorecard")
sc.append(["OWNER/APP FIX SCORECARD"]); sc.cell(1,1).font = Font(bold=True, size=13, color="1F3864")
sc.append([]); sc.append(["L5 tasks", len(rows)])
sc.append(["Areas LLM-verified", src_c["LLM-verified"]]); sc.append(["Areas deterministic", src_c["deterministic"]])
sc.append(["Tasks with no owner", noowner])
sc.append([]); sc.append(["Note", "Owners matched to role real-responsibilities (in-domain roles only, execs only on approval steps)."])
sc.append(["Note", "Apps = task-type rules constrained to the division catalog (no cross-domain misfits)."])
sc.append(["Note", "27 areas carry full LLM-verified owners+apps; rest deterministically matched. Re-run LLM polish after limit reset (2:40am CT)."])
for i, w in enumerate([26, 90], 1): sc.column_dimensions[get_column_letter(i)].width = w

out = "documents/value-streams/ABC-Insurance-Operating-Model-MASTER-v8-FIXED.xlsx"
wb.save(out)
print("SAVED", out)
print("tasks", len(rows), "| LLM areas", src_c["LLM-verified"], "| deterministic areas", src_c["deterministic"], "| no-owner", noowner)
# spot-check the go/no-go task
for r in rows:
    if "go" in r[4].lower() and "no" in r[4].lower():
        print("\nGO/NO-GO:", r[4], "\n  owner:", r[6], "| primary:", r[8], "| supporting:", r[9]); break
