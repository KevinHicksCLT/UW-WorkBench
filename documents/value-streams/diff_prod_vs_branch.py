# Forensic diff: PROD (ep-rough-bar / br-curly-poetry) vs BRANCH (ep-raspy-boat / proud-lab).
# Two lenses:
#   1. Row counts for every table.
#   2. SEMANTIC diff of key tables by NATURAL KEY (not cuid) with timestamps &
#      whitespace normalized out, so re-seed / formatting noise is not flagged as change.
import os, re, json
from collections import defaultdict
import psycopg
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
ENV  = os.path.join(HERE, "..", "..", "backend", ".env")
OUT  = os.path.join(HERE, "DB-Diff_prod-vs-branch.xlsx")

def branch_url():
    for line in open(ENV, encoding="utf-8"):
        if line.strip().startswith("DATABASE_URL="):
            return line.split("=",1)[1].strip().strip('"')
    raise SystemExit("no DATABASE_URL")
BRANCH_URL = branch_url()
PROD_URL   = BRANCH_URL.replace("ep-raspy-boat-aq1uy5u0", "ep-rough-bar-aqp1cb8v")
assert "ep-rough-bar" in PROD_URL, "prod host swap failed"

prod = psycopg.connect(PROD_URL); pcur = prod.cursor()
brch = psycopg.connect(BRANCH_URL); bcur = brch.cursor()

# ── styling ────────────────────────────────────────────────────────────────
HF=PatternFill("solid",fgColor="1F3864"); HFONT=Font(bold=True,color="FFFFFF",size=10)
SUB=Font(bold=True,size=12,color="1F3864")
ADD=PatternFill("solid",fgColor="C6EFCE"); REM=PatternFill("solid",fgColor="FFC7CE")
CHG=PatternFill("solid",fgColor="FFEB9C"); SAME=PatternFill("solid",fgColor="FFFFFF")
TH=Side(style="thin",color="BFBFBF"); BD=Border(left=TH,right=TH,top=TH,bottom=TH)
def hdr(ws,row,n):
    for c in range(1,n+1):
        x=ws.cell(row,c); x.fill=HF; x.font=HFONT; x.border=BD
        x.alignment=Alignment(vertical="center",wrap_text=True)
    ws.row_dimensions[row].height=26
def widths(ws,ws_widths):
    for i,w in enumerate(ws_widths,1): ws.column_dimensions[get_column_letter(i)].width=w

wb=openpyxl.Workbook(); wb.remove(wb.active)

# ── lens 1: row counts for every table ──────────────────────────────────────
pcur.execute("""SELECT table_name FROM information_schema.tables
                WHERE table_schema='public' AND table_type='BASE TABLE'
                AND table_name<>'_prisma_migrations' ORDER BY table_name""")
tables=[r[0] for r in pcur.fetchall()]
def count(cur,t):
    cur.execute(f'SELECT count(*) FROM "{t}"'); return cur.fetchone()[0]

ws=wb.create_sheet("Table Row Counts")
ws.cell(1,1,"Row counts — PROD (ep-rough-bar) vs BRANCH (proud-lab/ep-raspy-boat)").font=Font(bold=True,size=14,color="1F3864")
ws.cell(2,1,"Δ = branch − prod. Green = branch has more, Red = branch has fewer, blank = identical.").font=Font(italic=True,size=9,color="555555")
head=["Table","PROD rows","BRANCH rows","Δ (branch−prod)","Status"]
for c,h in enumerate(head,1): ws.cell(4,c,h)
hdr(ws,4,len(head))
r=5
count_rows=[]
for t in tables:
    p=count(pcur,t); b=count(bcur,t); d=b-p
    count_rows.append((t,p,b,d))
for t,p,b,d in count_rows:
    status="identical" if d==0 and p==b else ("branch +" if d>0 else "branch -")
    for c,v in enumerate([t,p,b,d,status],1):
        x=ws.cell(r,c,v); x.border=BD
    if d>0: ws.cell(r,5).fill=ADD
    elif d<0: ws.cell(r,5).fill=REM
    r+=1
ws.freeze_panes="A5"; widths(ws,[34,14,14,16,14])
print("counted", len(tables), "tables")

# ── lens 2: semantic diff by natural key ─────────────────────────────────────
def norm(v):
    if v is None: return ""
    if isinstance(v,bool): return "1" if v else "0"
    s=str(v).strip().lower(); s=re.sub(r"\s+"," ",s); return s

SPECS=[
 {"name":"Division","key":["k"],"cmp":["code","hc"],
  "sql":'SELECT lower(trim(name)) k, code, "higherCategory" hc FROM "Division"'},
 {"name":"Role","key":["k"],"cmp":["rf","rl","div","dep","pd","pvs","st"],
  "sql":'''SELECT lower(trim(r.name)) k, r."roleFamily" rf, r."roleLevel" rl,
            lower(coalesce(d.name,'')) div, lower(coalesce(dep.name,'')) dep,
            r."primaryDomain" pd, r."primaryValueStream" pvs, coalesce(r.status,'') st
           FROM "Role" r LEFT JOIN "Division" d ON d.id=r."divisionId"
           LEFT JOIN "Department" dep ON dep.id=r."departmentId"'''},
 {"name":"Application","key":["k"],"cmp":["kind","category","vendor","criticality","sor","pdn","code"],
  "sql":'SELECT lower(trim(name)) k, kind, category, vendor, criticality, "systemOfRecord" sor, "primaryDivisionName" pdn, code FROM "Application"'},
 {"name":"ValueStream","key":["k"],"cmp":["dom"],
  "sql":'''SELECT lower(trim(name)) k, coalesce(domain,'') dom FROM "ValueStream"'''},
 {"name":"Level (value-stream tree)","key":["ln","nm","par"],"cmp":[],
  "sql":'''SELECT l."levelNumber" ln, lower(trim(l.name)) nm, lower(coalesce(p.name,'')) par
           FROM "Level" l LEFT JOIN "Level" p ON p.id=l."parentId"'''},
 {"name":"ProcessStep","key":["vs","l3","l4","nm"],"cmp":[],
  "sql":'''SELECT lower(coalesce(v.name,'')) vs, lower(coalesce(ps.l3,'')) l3,
            lower(coalesce(ps.l4,'')) l4, lower(trim(ps.name)) nm
           FROM "ProcessStep" ps LEFT JOIN "ValueStream" v ON v.id=ps."valueStreamId"'''},
 {"name":"Deliverable","key":["t","o"],"cmp":["type","status"],
  "sql":'''SELECT lower(trim(title)) t, lower(coalesce(owner,'')) o, type, status FROM "Deliverable"'''},
 {"name":"Task","key":["t","d"],"cmp":["status","priority","source","ai"],
  "sql":'''SELECT lower(trim(t.title)) t, lower(coalesce(d.title,'')) d, t.status, t.priority, t.source, coalesce(t."aiDisposition",'') ai
           FROM "Task" t LEFT JOIN "Deliverable" d ON d.id=t."deliverableId"'''},
 {"name":"ChecklistItem","key":["role","txt"],"cmp":["cr"],
  "sql":'''SELECT lower(trim(r.name)) role, lower(trim(ci.text)) txt, ci."crossRole" cr
           FROM "ChecklistItem" ci JOIN "Role" r ON r.id=ci."roleId"'''},
 {"name":"RoleTask","key":["role","txt"],"cmp":[],
  "sql":'''SELECT lower(trim(r.name)) role, lower(trim(rt.text)) txt
           FROM "RoleTask" rt JOIN "Role" r ON r.id=rt."roleId"'''},
]

def load(cur,spec):
    cur.execute(spec["sql"]); cols=[d.name for d in cur.description]
    rows={}
    for row in cur.fetchall():
        rec=dict(zip(cols,row))
        k=tuple(norm(rec[c]) for c in spec["key"])
        rows[k]=tuple(norm(rec[c]) for c in spec["cmp"])
    return rows

# summary sheet
sw=wb.create_sheet("Semantic Diff Summary")
sw.cell(1,1,"Semantic diff by NATURAL KEY (cuid ignored; timestamps & whitespace normalized out)").font=Font(bold=True,size=14,color="1F3864")
sw.cell(2,1,"'Changed' = same natural key on both sides but a compared business field differs. Detail sheets list the rows.").font=Font(italic=True,size=9,color="555555")
sh=["Table","Natural key","PROD keys","BRANCH keys","Only in PROD (removed)","Only in BRANCH (added)","In both","Changed (semantic)"]
for c,h in enumerate(sh,1): sw.cell(4,c,h)
hdr(sw,4,len(sh)); sr=5
summary=[]
for spec in SPECS:
    P=load(pcur,spec); B=load(bcur,spec)
    pk=set(P); bk=set(B)
    only_p=pk-bk; only_b=bk-pk; both=pk&bk
    changed=[k for k in both if P[k]!=B[k]]
    summary.append((spec,P,B,only_p,only_b,both,changed))
    vals=[spec["name"],"+".join(spec["key"]),len(P),len(B),len(only_p),len(only_b),len(both),len(changed)]
    for c,v in enumerate(vals,1):
        x=sw.cell(sr,c,v); x.border=BD
    if len(only_p): sw.cell(sr,5).fill=REM
    if len(only_b): sw.cell(sr,6).fill=ADD
    if len(changed): sw.cell(sr,8).fill=CHG
    sr+=1
    print(f'{spec["name"]}: prod {len(P)} branch {len(B)} | -{len(only_p)} +{len(only_b)} ~{len(changed)}')
sw.freeze_panes="A5"; widths(sw,[26,22,12,13,22,22,12,18])

# detail sheets (cap big ones)
CAP=3000
DETAIL={"Division","Role","Application","ValueStream","Level (value-stream tree)","Deliverable","ChecklistItem"}
def build_detail(spec,P,B,only_p,only_b,changed):
    nm="Diff - "+spec["name"].split(" (")[0]
    d=wb.create_sheet(nm[:31])
    kcols=spec["key"]; ccols=spec["cmp"]
    head=["Change"]+kcols+[f"PROD.{c}" for c in ccols]+[f"BRANCH.{c}" for c in ccols]
    for c,h in enumerate(head,1): d.cell(1,c,h)
    hdr(d,1,len(head))
    state={"rr":2,"n":0}
    def emit(tag,k,pv,bv,fill):
        if state["n"]>=CAP: return
        rr=state["rr"]
        row=[tag]+list(k)+ (list(pv) if pv else [""]*len(ccols)) + (list(bv) if bv else [""]*len(ccols))
        for c,v in enumerate(row,1):
            x=d.cell(rr,c,v); x.border=BD; x.alignment=Alignment(vertical="top",wrap_text=True)
        d.cell(rr,1).fill=fill; state["rr"]+=1; state["n"]+=1
    for k in sorted(only_p): emit("REMOVED (prod only)",k,P[k],None,REM)
    for k in sorted(only_b): emit("ADDED (branch only)",k,None,B[k],ADD)
    for k in sorted(changed): emit("CHANGED",k,P[k],B[k],CHG)
    if state["n"]>=CAP: d.cell(state["rr"],1,f"… capped at {CAP} rows").font=Font(italic=True,color="FF0000")
    d.freeze_panes="A2"
    widths(d,[20]+[26]*len(kcols)+[18]*(2*len(ccols)))
for spec,P,B,only_p,only_b,both,changed in summary:
    if spec["name"] in DETAIL:
        build_detail(spec,P,B,only_p,only_b,changed)

prod.close(); brch.close()
wb.save(OUT)
print("SAVED:",OUT)
