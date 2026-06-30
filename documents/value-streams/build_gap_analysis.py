# Builds the gap-analysis + DB-listing sheets into the v007 workbook.
# Reads the active Neon DB (ep-raspy-boat / proud-lab) via backend/.env DATABASE_URL.
import os, re, sys, shutil
from collections import defaultdict
import psycopg
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(HERE, "ABC-Insurance-Operating-Model-REFINED-v007.xlsx")
ENV = os.path.join(HERE, "..", "..", "backend", ".env")

def db_url():
    for line in open(ENV, encoding="utf-8"):
        line = line.strip()
        if line.startswith("DATABASE_URL="):
            return line.split("=", 1)[1].strip().strip('"')
    raise SystemExit("no DATABASE_URL")

def norm(s):
    if s is None: return ""
    s = str(s).lower().strip()
    s = re.sub(r"[&/]", " and ", s)
    s = re.sub(r"[^a-z0-9 ]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s

# ── styling helpers ───────────────────────────────────────────────────────
HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
SUB_FILL = PatternFill("solid", fgColor="D6DCE5")
SUB_FONT = Font(bold=True, size=11, color="1F3864")
OK_FILL  = PatternFill("solid", fgColor="C6EFCE")
MISS_FILL= PatternFill("solid", fgColor="FFC7CE")
EXTRA_FILL=PatternFill("solid", fgColor="FFEB9C")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def style_header(ws, row, ncol):
    for c in range(1, ncol+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL; cell.font = HDR_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = 28

def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ── 1. read spreadsheet hierarchy ──────────────────────────────────────────
print("reading workbook…")
wb = openpyxl.load_workbook(SRC)
ws5 = wb["L5 Process List"]
rows = list(ws5.iter_rows(min_row=2, values_only=True))
# cols: L1,L2,L3,L3#,L4,L4#,L5,L5#,How,App,Checklist,Roles,Deliverables
ss_l2 = {}                       # (L1,L2) -> {l3set, l4set, l5count}
ss_l3 = {}                       # (L1,L2,L3) -> {l4set, l5count}
for r in rows:
    if not r or r[0] is None: continue
    l1,l2,l3,_,l4,_,l5 = (str(r[i]).strip() if r[i] is not None else "" for i in range(7))
    if not l2: continue
    k2=(l1,l2); k3=(l1,l2,l3)
    d2=ss_l2.setdefault(k2,{"l3":set(),"l4":set(),"l5":0})
    d2["l3"].add(l3); d2["l4"].add((l3,l4)); d2["l5"]+=1
    d3=ss_l3.setdefault(k3,{"l4":set(),"l5":0})
    d3["l4"].add(l4); d3["l5"]+=1
print(f"spreadsheet: {len(ss_l2)} divisions, {len(ss_l3)} L3 processes, {len(rows)} L5 rows")

# ── 2. read DB Level tree (canonical app value streams) ─────────────────────
print("querying DB…")
conn = psycopg.connect(db_url())
cur = conn.cursor()
cur.execute('SELECT id,"parentId","levelNumber",name FROM "Level"')
lv = {r[0]:{"parent":r[1],"n":r[2],"name":r[3]} for r in cur.fetchall()}

def path(nid):
    """return dict levelNumber->name up the tree"""
    out={}
    while nid and nid in lv:
        out[lv[nid]["n"]]=lv[nid]["name"]; nid=lv[nid]["parent"]
    return out

db_l2={}   # (L1,L2)->{l3,l4,l5count}
db_l3={}   # (L1,L2,L3)->{l4,l5count}
for nid,info in lv.items():
    p=path(nid)
    l1=p.get(1,""); l2=p.get(2,""); l3=p.get(3,""); l4=p.get(4,""); n=info["n"]
    if n==3:
        db_l2.setdefault((l1,l2),{"l3":set(),"l4":set(),"l5":0})["l3"].add(l3)
        db_l3.setdefault((l1,l2,l3),{"l4":set(),"l5":0})
    elif n==4:
        db_l2.setdefault((l1,l2),{"l3":set(),"l4":set(),"l5":0})["l4"].add((l3,l4))
        db_l3.setdefault((l1,l2,l3),{"l4":set(),"l5":0})["l4"].add(l4)
    elif n==5:
        db_l2.setdefault((l1,l2),{"l3":set(),"l4":set(),"l5":0})["l5"]+=1
        db_l3.setdefault((l1,l2,l3),{"l4":set(),"l5":0})["l5"]+=1

# legacy / overall counts
def scalar(q):
    cur.execute(q); return cur.fetchone()[0]
counts = {
 "ValueStream": scalar('SELECT count(*) FROM "ValueStream"'),
 "ProcessStep": scalar('SELECT count(*) FROM "ProcessStep"'),
 "Role": scalar('SELECT count(*) FROM "Role"'),
 "Application": scalar('SELECT count(*) FROM "Application"'),
 "ChecklistItem": scalar('SELECT count(*) FROM "ChecklistItem"'),
 "RoleTask": scalar('SELECT count(*) FROM "RoleTask"'),
 "Task": scalar('SELECT count(*) FROM "Task"'),
 "Deliverable": scalar('SELECT count(*) FROM "Deliverable"'),
}
lvl_counts={n:scalar(f'SELECT count(*) FROM "Level" WHERE "levelNumber"={n}') for n in range(6)}
print("level counts", lvl_counts)

# ── build name->division lookups (normalized) for matching ──────────────────
ss_div_by_norm={norm(l2):(l1,l2) for (l1,l2) in ss_l2}
db_div_by_norm={norm(l2):(l1,l2) for (l1,l2) in db_l2}
ss_l3_by_norm={(norm(l2),norm(l3)):(l1,l2,l3) for (l1,l2,l3) in ss_l3}
db_l3_by_norm={(norm(l2),norm(l3)):(l1,l2,l3) for (l1,l2,l3) in db_l3}

# ════════════════════════════════════════════════════════════════════════════
# SHEET: Value Stream Gap Analysis
# ════════════════════════════════════════════════════════════════════════════
for nm in ["VS Gap Analysis","DB - Roles","DB - Applications","DB - Checklist Items",
           "DB - Role Tasks","DB - Tasks (Tracker)","DB - Deliverables"]:
    if nm in wb.sheetnames: del wb[nm]

gw = wb.create_sheet("VS Gap Analysis", index=1)
r=1
gw.cell(r,1,"Value-Stream Gap Analysis — Spreadsheet v007  vs  Application/Database").font=Font(bold=True,size=14,color="1F3864"); r+=1
gw.cell(r,1,"Compares the v007 operating-model hierarchy against the canonical Level tree that drives the app's Value Streams tab. "
            "The Level tree was rebuilt from this workbook in 'REFINED-v2 (phase 1)' and is a PARTIAL load — this sheet quantifies the remaining gap.").font=Font(italic=True,size=9,color="555555"); r+=2

# Summary block
gw.cell(r,1,"1 · Level-by-level counts").font=SUB_FONT; r+=1
hdr=["Level","Spreadsheet v007","DB — Level tree (app value streams)","DB — legacy tables","Gap (SS − Level tree)"]
for c,h in enumerate(hdr,1): gw.cell(r,c,h)
style_header(gw,r,len(hdr)); r+=1
ss_levels={
 "L1 Segment": len({l1 for (l1,l2) in ss_l2}),
 "L2 Division": len(ss_l2),
 "L3 Process / Value Stream": len(ss_l3),
 "L4 Sub-Process": sum(len(d["l4"]) for d in ss_l2.values()),
 "L5 Step": sum(d["l5"] for d in ss_l2.values()),
}
db_levels={
 "L1 Segment":lvl_counts[1],"L2 Division":lvl_counts[2],"L3 Process / Value Stream":lvl_counts[3],
 "L4 Sub-Process":lvl_counts[4],"L5 Step":lvl_counts[5],
}
legacy={"L1 Segment":"—","L2 Division":"—","L3 Process / Value Stream":f'{counts["ValueStream"]} (ValueStream)',
        "L4 Sub-Process":"—","L5 Step":f'{counts["ProcessStep"]} (ProcessStep)'}
for k in ss_levels:
    gap=ss_levels[k]-db_levels[k]
    vals=[k,ss_levels[k],db_levels[k],legacy[k],gap]
    for c,v in enumerate(vals,1):
        cell=gw.cell(r,c,v); cell.border=BORDER
        if c==5 and isinstance(gap,int) and gap>0: cell.fill=MISS_FILL
        if c==5 and gap==0: cell.fill=OK_FILL
    r+=1
r+=1

# L2 Division gap
gw.cell(r,1,"2 · Division (L2) coverage").font=SUB_FONT; r+=1
hdr=["L1 Segment","L2 Division","In spreadsheet?","In app/DB?","SS L3","DB L3","SS L4","DB L4","SS L5","DB L5","Status"]
for c,h in enumerate(hdr,1): gw.cell(r,c,h)
style_header(gw,r,len(hdr)); r+=1
all_div=set(ss_div_by_norm)|set(db_div_by_norm)
def div_rows():
    out=[]
    for nrm in sorted(all_div):
        sk=ss_div_by_norm.get(nrm); dk=db_div_by_norm.get(nrm)
        l1=(sk or dk)[0]; l2=(sk or dk)[1]
        sd=ss_l2.get(sk) if sk else None; dd=db_l2.get(dk) if dk else None
        in_ss="Yes" if sk else "No"; in_db="Yes" if dk else "No"
        status="✔ Match" if sk and dk else ("✖ Missing in app/DB" if sk else "➕ Extra in app/DB")
        out.append((l1,l2,in_ss,in_db,
            len(sd["l3"]) if sd else 0, len(dd["l3"]) if dd else 0,
            len(sd["l4"]) if sd else 0, len(dd["l4"]) if dd else 0,
            sd["l5"] if sd else 0, dd["l5"] if dd else 0, status))
    return out
for row in sorted(div_rows(), key=lambda x:(x[0],x[1])):
    for c,v in enumerate(row,1):
        cell=gw.cell(r,c,v); cell.border=BORDER
        if c==11:
            cell.fill=OK_FILL if v.startswith("✔") else (MISS_FILL if v.startswith("✖") else EXTRA_FILL)
    r+=1
r+=1

# L3 process gap
gw.cell(r,1,"3 · Process / Value-Stream (L3) coverage").font=SUB_FONT; r+=1
hdr=["L1 Segment","L2 Division","L3 Process / Value Stream","In spreadsheet?","In app/DB?","SS L4","DB L4","SS L5","DB L5","Status"]
for c,h in enumerate(hdr,1): gw.cell(r,c,h)
style_header(gw,r,len(hdr)); r+=1
all_l3=set(ss_l3_by_norm)|set(db_l3_by_norm)
l3rows=[]
for key in all_l3:
    sk=ss_l3_by_norm.get(key); dk=db_l3_by_norm.get(key)
    l1,l2,l3=(sk or dk)
    sd=ss_l3.get(sk) if sk else None; dd=db_l3.get(dk) if dk else None
    status="✔ Match" if sk and dk else ("✖ Missing in app/DB" if sk else "➕ Extra in app/DB")
    l3rows.append((l1,l2,l3,"Yes" if sk else "No","Yes" if dk else "No",
        len(sd["l4"]) if sd else 0, len(dd["l4"]) if dd else 0,
        sd["l5"] if sd else 0, dd["l5"] if dd else 0, status))
for row in sorted(l3rows, key=lambda x:(x[0],x[1],x[2])):
    for c,v in enumerate(row,1):
        cell=gw.cell(r,c,v); cell.border=BORDER
        if c==10:
            cell.fill=OK_FILL if v.startswith("✔") else (MISS_FILL if v.startswith("✖") else EXTRA_FILL)
    r+=1
gw.freeze_panes="A1"
autosize(gw,[16,34,30,14,12,8,8,8,8,8,22])

# ════════════════════════════════════════════════════════════════════════════
# DB listing sheets
# ════════════════════════════════════════════════════════════════════════════
def dump(sheet, query, headers, widths):
    cur.execute(query)
    data=cur.fetchall()
    ws=wb.create_sheet(sheet)
    ws.cell(1,1,f"{sheet}  —  {len(data)} rows  (source: active Neon DB, proud-lab/ep-raspy-boat)").font=Font(bold=True,size=11,color="1F3864")
    for c,h in enumerate(headers,1): ws.cell(2,c,h)
    style_header(ws,2,len(headers))
    for ri,row in enumerate(data,start=3):
        for c,v in enumerate(row,1):
            cell=ws.cell(ri,c,v); cell.border=BORDER
            cell.alignment=Alignment(vertical="top",wrap_text=True)
    ws.freeze_panes="A3"; autosize(ws,widths)
    print(f"{sheet}: {len(data)} rows")

dump("DB - Roles",
 '''SELECT r.name, r."roleFamily", r."roleLevel", d.name, dep.name, r."primaryDomain",
           r."primaryValueStream", r.status, m.name
    FROM "Role" r
    LEFT JOIN "Division" d ON d.id=r."divisionId"
    LEFT JOIN "Department" dep ON dep.id=r."departmentId"
    LEFT JOIN "Role" m ON m.id=r."managerRoleId"
    ORDER BY r."roleFamily", r.name''',
 ["Role","Role Family","Role Level","Division","Department","Primary Domain","Primary Value Stream","Status","Reports To"],
 [40,24,20,26,26,22,28,12,32])

dump("DB - Applications",
 '''SELECT name, kind, category, vendor, criticality, "ownershipModel",
           "systemOfRecord", code, description
    FROM "Application" ORDER BY category NULLS LAST, name''',
 ["Application","Kind","Category","Vendor","Criticality","Ownership","System of Record","Code","Description"],
 [34,12,16,22,12,14,16,10,50])

dump("DB - Checklist Items",
 '''SELECT r.name, c.name, ci.text, ci."crossRole", ci."sourceSheet"
    FROM "ChecklistItem" ci
    JOIN "Role" r ON r.id=ci."roleId"
    LEFT JOIN "Category" c ON c.id=ci."categoryId"
    ORDER BY r.name, c.name NULLS LAST''',
 ["Role","Category","Checklist Item","Cross-Role","Source Sheet"],
 [38,24,90,12,18])

dump("DB - Role Tasks",
 '''SELECT r.name, c.name, rt.text, rt."sourceSheet"
    FROM "RoleTask" rt
    JOIN "Role" r ON r.id=rt."roleId"
    LEFT JOIN "Category" c ON c.id=rt."categoryId"
    ORDER BY r.name, c.name NULLS LAST''',
 ["Role","Category","Task","Source Sheet"],
 [38,24,90,18])

dump("DB - Tasks (Tracker)",
 '''SELECT t.title, d.title, t.owner, t.status, t.priority, t.source,
           t."aiDisposition", t."agentScore"
    FROM "Task" t LEFT JOIN "Deliverable" d ON d.id=t."deliverableId"
    ORDER BY d.title NULLS LAST, t.title''',
 ["Task","Deliverable","Owner","Status","Priority","Source","AI Disposition","Agent Score"],
 [70,40,24,14,10,12,16,12])

dump("DB - Deliverables",
 '''SELECT title, type, owner, status, "dueDate", description
    FROM "Deliverable" ORDER BY type, title''',
 ["Deliverable","Type","Owner","Status","Due Date","Description"],
 [50,14,24,14,16,60])

# ── save (avoid clobbering an Excel-locked original) ───────────────────────
out=SRC
try:
    wb.save(out)
    print("SAVED in place:", out)
except PermissionError:
    out=SRC.replace(".xlsx","-with-gap-analysis.xlsx")
    wb.save(out)
    print("ORIGINAL LOCKED — saved new file:", out)
conn.close()
