# Associates EXISTING DB data (division-grain) onto the L5 Process List rows.
# No new data is invented. Fuzzy division mappings are highlighted yellow.
import os, re
from collections import defaultdict
import psycopg
import openpyxl
from openpyxl.styles import PatternFill, Font

HERE = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(HERE, "ABC-Insurance-Operating-Model-REFINED-v007.xlsx")
ENV = os.path.join(HERE, "..", "..", "backend", ".env")
YELLOW = PatternFill("solid", fgColor="FFEB9C")
CELL_MAX = 32000   # < Excel hard limit 32767

def db_url():
    for line in open(ENV, encoding="utf-8"):
        if line.strip().startswith("DATABASE_URL="):
            return line.split("=", 1)[1].strip().strip('"')
    raise SystemExit("no DATABASE_URL")

# ── SS division → (DB division, confident?) ────────────────────────────────
MAP = {
 "Actuarial": ("Actuarial", True),
 "Claims": ("Claims", True),
 "Reinsurance": ("Reinsurance", True),
 "Sales, Distribution & Marketing": ("Sales, Distribution & Marketing", True),
 "Underwriting": ("Underwriting", True),
 "Finance & Investments": ("Finance & Investments", True),
 "Human Resources & Talent": ("Human Resources & Talent", True),
 "Legal & Corporate Governance": ("Legal & Corporate Governance", True),
 "Risk, Compliance & Audit": ("Risk, Compliance & Audit", True),
 "Cybersecurity & IAM": ("Cybersecurity & IAM", True),
 "Data & AI": ("Data & AI", True),
 "Technology & Engineering": ("Technology & Engineering", True),
 # fuzzy merges (DB consolidated these) → highlight yellow
 "Business Operations": ("Operations & Customer Service", False),
 "Call Center": ("Operations & Customer Service", False),
 "Corporate Functions Operations": ("Operations & Customer Service", False),
 "Product & Delivery": ("Product, Delivery & PMO", False),
 "Program Management Office": ("Product, Delivery & PMO", False),
}

conn = psycopg.connect(db_url()); cur = conn.cursor()

# roles per DB division
cur.execute('''SELECT d.name, r.name FROM "Division" d JOIN "Role" r ON r."divisionId"=d.id
               ORDER BY d.name, r.name''')
roles = defaultdict(list)
for dn, rn in cur.fetchall():
    if rn not in roles[dn]: roles[dn].append(rn)

# apps per DB division (primaryDivisionName)
cur.execute('''SELECT "primaryDivisionName", name, COALESCE("systemOfRecord",false)
               FROM "Application" WHERE "primaryDivisionName" IS NOT NULL ORDER BY 1,2''')
apps = defaultdict(list)
for dn, an, sor in cur.fetchall():
    apps[dn].append(f"{an} [SoR]" if sor else an)

# deliverables per DB division (owner text → Role.name → division)
cur.execute('''SELECT d.name, dl.title
               FROM "Deliverable" dl
               JOIN "Role" r ON r.name = dl.owner
               JOIN "Division" d ON d.id = r."divisionId"
               WHERE dl.owner IS NOT NULL ORDER BY d.name, dl.title''')
delivs = defaultdict(list)
for dn, t in cur.fetchall():
    if t not in delivs[dn]: delivs[dn].append(t)
cur.execute('SELECT count(*) FROM "Deliverable" dl LEFT JOIN "Role" r ON r.name=dl.owner WHERE r.id IS NULL')
deliv_unmatched = cur.fetchone()[0]

# checklist counts per DB division
cur.execute('''SELECT d.name, count(ci.id)
               FROM "Division" d JOIN "Role" r ON r."divisionId"=d.id
               JOIN "ChecklistItem" ci ON ci."roleId"=r.id GROUP BY d.name''')
chk = {dn: n for dn, n in cur.fetchall()}
conn.close()

def join_capped(items):
    if not items: return ""
    s = "; ".join(items)
    if len(s) <= CELL_MAX: return s
    out=[]; n=0
    for it in items:
        if n + len(it) + 2 > CELL_MAX - 40: break
        out.append(it); n += len(it)+2
    return "; ".join(out) + f"  …(+{len(items)-len(out)} more)"

# ── write into L5 Process List (cols: 10 App, 11 Checklist, 12 Roles, 13 Deliv) ──
print("loading workbook…")
wb = openpyxl.load_workbook(SRC)
ws = wb["L5 Process List"]
stats = defaultdict(int); filled = 0; yellowed = 0; unmapped = set()
for row in ws.iter_rows(min_row=2):
    l2 = row[1].value
    if l2 is None: continue
    l2 = str(l2).strip()
    if l2 not in MAP:
        unmapped.add(l2); continue
    dbdiv, confident = MAP[l2]
    role_s  = join_capped(roles.get(dbdiv, []))
    app_s   = join_capped(apps.get(dbdiv, []))
    deliv_s = join_capped(delivs.get(dbdiv, []))
    n_chk   = chk.get(dbdiv, 0)
    chk_s   = f"{n_chk} checklist items — see 'DB - Checklist Items' (role-level)" if n_chk else ""
    row[9].value, row[10].value, row[11].value, row[12].value = app_s, chk_s, role_s, deliv_s
    if not confident:
        for ci in (9, 10, 11, 12):
            row[ci].fill = YELLOW
        yellowed += 1
    filled += 1
    stats[l2] += 1

print(f"filled rows: {filled}, yellow rows: {yellowed}")
print("unmapped SS divisions (left blank):", sorted(unmapped) or "none")
print(f"deliverables with no role-owner match (dropped): {deliv_unmatched}")
try:
    wb.save(SRC)
    print("SAVED in place:", SRC)
except PermissionError:
    alt = SRC.replace(".xlsx", "-DB-associated.xlsx")
    wb.save(alt)
    print("ORIGINAL LOCKED (close Excel) — saved new file:", alt)
