# Diff: BRANCH DB (proud-lab/ep-raspy-boat) vs the v7_062226 spreadsheet.
# v7 hierarchy == v007 (17 div / 135 L3 / 867 L4 / 8355 L5); the NEW content is the
# populated Roles / Deliverables / Application columns. So we reconcile the ENTITIES
# the sheet references against what actually exists in the DB.
import os, re
from collections import defaultdict, Counter
import psycopg
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE=os.path.dirname(os.path.abspath(__file__))
SHEET=r"C:\Users\xando\Downloads\ABC-Insurance-Operating-Model-v7_062226.xlsx"
ENV=os.path.join(HERE,"..","..","backend",".env")
OUT=os.path.join(HERE,"DB-Diff_branch-vs-v7sheet.xlsx")

def db_url():
    for line in open(ENV,encoding="utf-8"):
        if line.strip().startswith("DATABASE_URL="):
            return line.split("=",1)[1].strip().strip('"')
    raise SystemExit("no DATABASE_URL")

def norm(s):
    if s is None: return ""
    s=str(s).lower().strip()
    s=re.sub(r"\[[a-z]\]","",s)               # drop [C]/[R] tags
    s=re.sub(r"[&/]"," and ",s)
    s=re.sub(r"[^a-z0-9 ]"," ",s)
    s=re.sub(r"\s+"," ",s).strip()
    return s

# ── styling ──
HF=PatternFill("solid",fgColor="1F3864"); HFONT=Font(bold=True,color="FFFFFF",size=10)
ADD=PatternFill("solid",fgColor="C6EFCE"); REM=PatternFill("solid",fgColor="FFC7CE"); CHG=PatternFill("solid",fgColor="FFEB9C")
TH=Side(style="thin",color="BFBFBF"); BD=Border(left=TH,right=TH,top=TH,bottom=TH)
def hdr(ws,row,n):
    for c in range(1,n+1):
        x=ws.cell(row,c); x.fill=HF; x.font=HFONT; x.border=BD
        x.alignment=Alignment(vertical="center",wrap_text=True)
    ws.row_dimensions[row].height=26
def widths(ws,w):
    for i,x in enumerate(w,1): ws.column_dimensions[get_column_letter(i)].width=x

# ── parse spreadsheet ──
print("reading spreadsheet…")
wb_in=openpyxl.load_workbook(SHEET, read_only=True)
ws=wb_in["L5 Process List"]
role_ref=Counter()        # norm role -> rows referencing
role_disp={}              # norm -> display
deliv_ref=Counter(); deliv_disp={}
app_text_rows=0
sheet_div=Counter()
RX_NUM=re.compile(r"^\(\d+\)\s*(.+?)\s*$")
RX_SUB=re.compile(r"^\s*[-•]\s*(.+?)\s*$")
k_blob=[]
for r in ws.iter_rows(min_row=2,values_only=True):
    if r[1] is None: continue
    sheet_div[str(r[1]).strip()]+=1
    O=r[14]; P=r[15]; K=r[10]
    if O:
        for ln in str(O).splitlines():
            m=RX_NUM.match(ln.strip())
            if m:
                nm=re.sub(r"\[[A-Za-z]\]$","",m.group(1)).strip()
                n=norm(nm)
                if n: role_ref[n]+=1; role_disp.setdefault(n,nm)
    if P:
        for ln in str(P).splitlines():
            s=ln.strip()
            m=RX_NUM.match(s)
            if m:
                nm=m.group(1).strip()
                n=norm(nm)
                if n: deliv_ref[n]+=1; deliv_disp.setdefault(n,nm)
            else:
                m2=RX_SUB.match(s)
                if m2:                      # sub-bullet = role
                    nm=re.sub(r"\[[A-Za-z]\]$","",m2.group(1)).strip()
                    n=norm(nm)
                    if n: role_ref[n]+=1; role_disp.setdefault(n,nm)
    if K:
        app_text_rows+=1; k_blob.append(str(K).lower())
print(f"sheet: {len(role_ref)} distinct roles, {len(deliv_ref)} distinct deliverables referenced; K populated rows={app_text_rows}")

# ── DB entities (branch) ──
print("querying branch DB…")
conn=psycopg.connect(db_url()); cur=conn.cursor()
cur.execute('SELECT name FROM "Role"'); db_roles={norm(r[0]):r[0] for r in cur.fetchall()}
cur.execute('SELECT DISTINCT title FROM "Deliverable"'); db_delivs={norm(r[0]):r[0] for r in cur.fetchall()}
cur.execute('SELECT name FROM "Application"'); db_apps=[r[0] for r in cur.fetchall()]
# hierarchy
cur.execute('SELECT "levelNumber",count(*) FROM "Level" GROUP BY 1'); lvl={n:c for n,c in cur.fetchall()}
cur.execute('SELECT count(*) FROM "ProcessStep"'); ps=cur.fetchone()[0]
conn.close()

wbk=openpyxl.Workbook(); wbk.remove(wbk.active)

# ── Sheet 0: hierarchy + summary ──
s0=wbk.create_sheet("Summary")
s0.cell(1,1,"Diff — BRANCH DB (proud-lab) vs spreadsheet v7_062226").font=Font(bold=True,size=14,color="1F3864")
rows=[
 ("",""),
 ("HIERARCHY","sheet vs DB"),
 ("L2 Divisions (sheet)", len(sheet_div)),
 ("L3 (sheet)", 135), ("L4 (sheet)", 867), ("L5 steps (sheet)", sum(sheet_div.values())),
 ("DB Level tree L2/L3/L4/L5", f"{lvl.get(2,0)} / {lvl.get(3,0)} / {lvl.get(4,0)} / {lvl.get(5,0)}"),
 ("DB ProcessStep rows", ps),
 ("",""),
 ("ENTITY RECONCILIATION","matched / sheet-only / DB-only"),
 ("Roles", f"{len(set(role_ref)&set(db_roles))} / {len(set(role_ref)-set(db_roles))} / {len(set(db_roles)-set(role_ref))}"),
 ("Deliverables", f"{len(set(deliv_ref)&set(db_delivs))} / {len(set(deliv_ref)-set(db_delivs))} / {len(set(db_delivs)-set(deliv_ref))}"),
 ("Distinct roles referenced in sheet", len(role_ref)),
 ("Distinct roles in DB", len(db_roles)),
 ("Distinct deliverables referenced in sheet", len(deliv_ref)),
 ("Distinct deliverables in DB", len(db_delivs)),
]
rr=3
for a,b in rows:
    s0.cell(rr,1,a); s0.cell(rr,2,b)
    if a in ("HIERARCHY","ENTITY RECONCILIATION"): s0.cell(rr,1).font=Font(bold=True,size=12,color="1F3864")
    rr+=1
widths(s0,[40,40])

# ── reconciliation sheet builder ──
def recon(title, ref, disp, dbmap, widths_):
    sh=wbk.create_sheet(title)
    sh.cell(1,1,title).font=Font(bold=True,size=12,color="1F3864")
    head=["Status","Name","Sheet refs (rows)","Matched DB name"]
    for c,h in enumerate(head,1): sh.cell(2,c,h)
    hdr(sh,2,4); r=3
    matched=sorted(set(ref)&set(dbmap)); sonly=sorted(set(ref)-set(dbmap)); donly=sorted(set(dbmap)-set(ref))
    for k in matched:
        for c,v in enumerate(["MATCH",disp.get(k,k),ref[k],dbmap[k]],1): sh.cell(r,c,v).border=BD
        r+=1
    for k in sonly:
        for c,v in enumerate(["SHEET ONLY (not in DB)",disp.get(k,k),ref[k],""],1):
            x=sh.cell(r,c,v); x.border=BD; x.fill=REM
        r+=1
    for k in donly:
        for c,v in enumerate(["DB ONLY (never referenced)",dbmap[k],0,dbmap[k]],1):
            x=sh.cell(r,c,v); x.border=BD; x.fill=ADD
        r+=1
    sh.freeze_panes="A3"; widths(sh,widths_)
    return len(matched),len(sonly),len(donly)

rm=recon("Roles Reconciliation", role_ref, role_disp, db_roles, [26,46,16,40])
dm=recon("Deliverables Reconciliation", deliv_ref, deliv_disp, db_delivs, [26,60,16,50])

# ── Apps: substring presence of each DB app in the K column text ──
sa=wbk.create_sheet("Applications Referenced")
sa.cell(1,1,"DB applications mentioned in the sheet's K (App/SOR) column").font=Font(bold=True,size=12,color="1F3864")
head=["DB Application","Mentioned in sheet?","≈ rows mentioning"]
for c,h in enumerate(head,1): sa.cell(2,c,h)
hdr(sa,2,3); r=3
blob="\n".join(k_blob)
for an in sorted(db_apps):
    nl=an.lower()
    cnt=sum(1 for k in k_blob if nl in k)
    for c,v in enumerate([an,"Yes" if cnt else "No",cnt],1):
        x=sa.cell(r,c,v); x.border=BD
        if c==2: x.fill=ADD if cnt else REM
    r+=1
sa.freeze_panes="A3"; widths(sa,[40,18,16])

wbk.save(OUT)
print("Roles  match/sheet-only/db-only:",rm)
print("Delivs match/sheet-only/db-only:",dm)
print("SAVED:",OUT)
